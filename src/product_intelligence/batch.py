from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from .input_identity import parse_product_query
from .discovery import search_web, search_web_for_fields
from .document_discovery import discover_product_documents
from .document_ingestion import process_pdf_document
from .excel_intake import analyze_workbook_intake
from .excel_mapper_v8 import fill_excel_v8
from .models import ProductIdentity, ProductRecord
from .normalize import key_norm
from .pdf_search_trace import PdfSearchTrace, format_trace_lines
from .pipeline import ProductPipeline
from .product_classification import classify_product
from .product_evidence_orchestrator import ProductEvidenceOrchestrator
from .record_builder import build_record_strict
from .resolution_engine import analyze_resolution
from .semantic_guard import is_placeholder
from .source_strategy import SourceStrategy
from .template_contract import analyze_template_contract
from .template_intelligence import analyze_matrix
from .universal_resolution_policy import ResolutionBudget, SearchBudgetTracker


RUNTIME_RESOLUTION_BUDGET = ResolutionBudget()
MAX_VALIDATED_SOURCES_PER_PRODUCT = RUNTIME_RESOLUTION_BUDGET.max_sources_accepted_per_product
SEARCH_STAGE_QUERY_QUOTAS = {
    "INITIAL": 4,
    "IDENTITY_REFINEMENT": 2,
    "MISSING_FIELDS": 2,
}


@dataclass
class BatchItem:
    row: int
    sheet: str
    identity: ProductIdentity
    source_url: str | None = None
    source_urls: list[str] | None = None


_TEMPLATE_IDENTITY_EXAMPLES = {"1234567890", "99999999", "999999999", "abc-1000-202", "abc1000202"}


def _clean_id(value):
    if value is None:return None
    text=str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():text=text[:-2]
    return text or None


def _identity_placeholder(value) -> bool:
    text=_clean_id(value)
    if not text:return True
    compact=key_norm(text).replace(" ","")
    if compact in _TEMPLATE_IDENTITY_EXAMPLES:return True
    return is_placeholder(text)


def _looks_like_part_number(value: str | None) -> bool:
    text=str(value or "").strip()
    if not text or " " in text or len(text)<6 or len(text)>48:return False
    if not re.fullmatch(r"[A-Za-z0-9._/-]+",text):return False
    if text.isdigit():return False
    return sum(ch.isalpha() for ch in text)>=2 and sum(ch.isdigit() for ch in text)>=1


def _promote_mpn(vals: dict[str, str]) -> None:
    if vals.get("mpn"):return
    for key in ("model","product_name"):
        value=vals.get(key)
        if _looks_like_part_number(value):vals["mpn"]=value;return


def detect_items(template: str) -> list[BatchItem]:
    intake=analyze_workbook_intake(template)
    return [BatchItem(row=p.row,sheet=p.sheet,identity=p.identity,source_url=p.source_url,source_urls=list(p.source_urls or [])) for p in intake.products]


def _best_product_sheet(template: str) -> tuple[str, int]:
    wb=load_workbook(template,data_only=False,read_only=False);best=None
    for ws in wb.worksheets:
        matrix=[[ws.cell(r,c).value for c in range(1,ws.max_column+1)] for r in range(1,min(ws.max_row,20)+1)]
        info=analyze_matrix(matrix);score=len(info.get("fields") or [])
        if score and (best is None or score>best[0]):best=(score,ws.title,info["header_row"])
    if not best:raise ValueError("No se pudo detectar la hoja de carga de productos del Excel.")
    return best[1],best[2]


def manual_identity_items(template: str, identities: list[ProductIdentity], source_urls_by_index: list[list[str]] | None = None) -> list[BatchItem]:
    sheet,header_row=_best_product_sheet(template);items=[]
    for i,ident in enumerate(identities):
        urls=[]
        if source_urls_by_index and i<len(source_urls_by_index):urls=list(dict.fromkeys(u for u in (source_urls_by_index[i] or []) if u))
        items.append(BatchItem(header_row+i+1,sheet,ident,source_urls=urls))
    return items


def manual_items(template: str, part_numbers: list[str]) -> list[BatchItem]:
    identities=[]
    for value in part_numbers:
        ident=parse_product_query(str(value))
        if ident:identities.append(ident)
    return manual_identity_items(template,identities)


def _meaningful_product_tokens(value: str | None) -> set[str]:
    stop={"the","and","with","for","of","de","del","con","para","headphone","headphones","headset","auricular","auriculares","wireless","wired","black","blue","white","negro","azul","blanco","on","ear","in"}
    return {x for x in re.split(r"[^a-z0-9]+",key_norm(value or "")) if len(x)>=2 and x not in stop}


def _cross_source_consistent(primary: ProductRecord, other: ProductRecord, url: str) -> bool:
    mpn=str(primary.identity.mpn or "").strip();compact_url=re.sub(r"[^a-z0-9]","",key_norm(url or ""));compact_mpn=re.sub(r"[^a-z0-9]","",key_norm(mpn))
    if compact_mpn and compact_mpn in compact_url:return True
    a=_meaningful_product_tokens(primary.identity.product_name or primary.identity.model);b=_meaningful_product_tokens(other.identity.product_name or other.identity.model)
    if not a or not b:return False
    shared=a&b
    return len(shared)>=max(2,min(3,len(a)//2))


def _merge_valid_records(records: list[ProductRecord]) -> ProductRecord:
    if not records:raise ValueError("no records to merge")
    def rank(rec:ProductRecord):return (2 if (rec.fetch or {}).get("source_class")=="manufacturer" else 1,2 if rec.identity.match_level=="EXACT" else 1,float(rec.identity.confidence or 0),len(rec.evidence))
    ordered=sorted(records,key=rank,reverse=True);primary=ordered[0];evidence=[];sources=[];warnings=[];notes=[];media_by_url={}
    for rec in ordered:
        evidence.extend(rec.evidence);sources.extend(rec.sources);warnings.extend(rec.warnings);notes.extend(rec.technical_notes)
        for item in rec.media:
            url=item.get("url")
            if not url:continue
            previous=media_by_url.get(url)
            if previous is None or item.get("confidence",0)>previous.get("confidence",0):media_by_url[url]=item
    merged=build_record_strict(primary.identity,evidence,list(dict.fromkeys(sources)));merged.media=list(media_by_url.values())
    merged.images=[m for m in merged.media if m.get("media_type")=="image" and m.get("scope") in {"EXACT_VARIANT","EXACT_PRODUCT"} and m.get("confidence",0)>=.80 and m.get("autofill_eligible")]
    merged.videos=[m for m in merged.media if m.get("media_type")=="video" and m.get("scope") in {"EXACT_VARIANT","EXACT_PRODUCT"} and m.get("confidence",0)>=.80 and m.get("autofill_eligible")]
    merged.warnings=list(dict.fromkeys(warnings));merged.technical_notes=notes;merged.site_profile=primary.site_profile
    merged.fetch={"method":"multi_source","source_class":(primary.fetch or {}).get("source_class"),"validated_sources":len(ordered),"manufacturer_sources":sum(1 for r in ordered if (r.fetch or {}).get("source_class")=="manufacturer"),"source_decisions":[(r.fetch or {}).get("source_decision") for r in ordered if (r.fetch or {}).get("source_decision")],"source_validation_counts":{"validated_sources":len(ordered),"material_evidence":len(merged.evidence or []),"final_specifications":len(merged.specifications or {})}}
    return merged


def _compact(value: str | None) -> str:return re.sub(r"[^a-z0-9]","",key_norm(value or ""))


def _candidate_official_domain(candidate, identity: ProductIdentity, accepted: list[ProductRecord]) -> str | None:
    host=(urlparse(candidate.url).hostname or "").lower().removeprefix("www.")
    if not host:return None
    if bool(getattr(candidate,"likely_official",False)):return host
    if getattr(candidate,"manual_source",False):
        label=_compact(host.split(".")[0]);strong=_compact(identity.mpn or identity.model or identity.product_name)
        if len(label)>=3 and strong and strong.startswith(label):return host
    return None


def _manufacturer_domains(records: list[ProductRecord]) -> set[str]:
    out=set()
    for rec in records:
        if (rec.fetch or {}).get("source_class")!="manufacturer":continue
        for url in [rec.fetch.get("final_url") if rec.fetch else None,*(rec.sources or [])]:
            if not url:continue
            host=(urlparse(url).hostname or "").lower().removeprefix("www.")
            if host:out.add(host)
    return out


def _resolution_for(records: list[ProductRecord], template_plan: dict | None) -> tuple[ProductRecord, dict]:
    merged=_merge_valid_records(records);return merged,analyze_resolution(merged,template_plan)


def _coverage_sufficient(resolution: dict, has_manufacturer: bool) -> bool:
    return bool(has_manufacturer and not resolution.get("blocked") and not list(resolution.get("research_terms") or []))


def _enriched_identity(item: BatchItem, rec: ProductRecord) -> ProductIdentity | None:
    learned_brand=rec.identity.brand or item.identity.brand;learned_model=rec.identity.model or rec.identity.product_name or item.identity.model or item.identity.product_name
    strong=item.identity.mpn or item.identity.ean or item.identity.upc or item.identity.gtin or rec.identity.mpn or rec.identity.gtin
    if not learned_brand or not strong:return None
    return ProductIdentity(mpn=item.identity.mpn or rec.identity.mpn,ean=item.identity.ean or rec.identity.ean,upc=item.identity.upc or rec.identity.upc,gtin=item.identity.gtin or rec.identity.gtin,brand=learned_brand,model=learned_model,confidence=max(float(item.identity.confidence or 0),float(rec.identity.confidence or 0)),match_level=rec.identity.match_level if rec.identity.match_level!="LOW" else item.identity.match_level,identifiers_confirmed=list(dict.fromkeys([*(item.identity.identifiers_confirmed or []),*(rec.identity.identifiers_confirmed or [])])))


def _prioritize(candidates) -> list:return sorted(candidates,key=lambda c:(not bool(getattr(c,"likely_official",False)),-float(getattr(c,"score",0))))
def _looks_like_pdf_url(url: str | None) -> bool:return (urlparse(str(url or "")).path or "").lower().endswith(".pdf")
def _product_key(identity: ProductIdentity) -> str:return str(identity.mpn or identity.ean or identity.upc or identity.gtin or identity.model or identity.product_name or "producto")


def _log_source_decision(rec: ProductRecord, log, prefix: str = "  ") -> bool:
    decision=(rec.fetch or {}).get("source_decision") or {};page_type=decision.get("page_type","UNKNOWN");identity=decision.get("identity",rec.identity.match_level or "UNKNOWN");authority=decision.get("authority",(rec.fetch or {}).get("source_class","unknown"));material=bool(decision.get("material_allowed",True));evidence_count=len(rec.evidence or [])
    allowed=material and identity in {"EXACT","COMPATIBLE","HIGH"} and evidence_count>0
    reason="OK" if allowed else ("PAGE_TYPE_NOT_MATERIAL" if not material else "IDENTITY_NOT_STRONG_ENOUGH" if identity not in {"EXACT","COMPATIBLE","HIGH"} else "NO_POLICY_APPROVED_EVIDENCE")
    log(f"{prefix}PAGE_TYPE={page_type} confidence={decision.get('page_type_confidence','?')}");log(f"{prefix}IDENTITY={identity} confidence={decision.get('identity_confidence','?')}");log(f"{prefix}AUTHORITY={authority} confidence={decision.get('authority_confidence','?')}");log(f"{prefix}EVIDENCE_ALLOWED={'YES' if allowed else 'NO'} reason={reason} evidence={evidence_count}")
    return allowed


def _log_source_rejection(exc: Exception, log, prefix: str = "  ") -> None:
    text=str(exc);marker="SOURCE_VALIDATION_REJECTED:"
    if marker in text:
        detail=text.split(marker,1)[1].strip();reason=detail.split()[0] if detail else "SOURCE_VALIDATION_REJECTED"
        if "identity=CONFLICT" in detail or "IDENTITY_CONFLICT" in detail:log(f"{prefix}IDENTITY=CONFLICT")
        if "page_type=" in detail:log(f"{prefix}PAGE_TYPE={detail.split('page_type=',1)[1].split()[0]}")
        if "authority=" in detail:log(f"{prefix}AUTHORITY={detail.split('authority=',1)[1].split()[0]}")
        log(f"{prefix}EVIDENCE_ALLOWED=NO reason={reason}")
    else:log(f"{prefix}SOURCE_REJECTED={type(exc).__name__}: {text}")


def _ingest_direct_documents(identity:ProductIdentity,*,target_semantics:list[str],seen_urls:set[str],errors:list[str],log,limit:int=6,budget_tracker:SearchBudgetTracker|None=None,accept_limit:int=3)->list[ProductRecord]:
    accepted=[];trace=PdfSearchTrace(_product_key(identity))
    pdf_limit=min(int(limit),budget_tracker.budget.max_pdfs_analyzed_per_product if budget_tracker else int(limit))
    document_candidates = discover_product_documents(identity,limit=pdf_limit,trace=trace)
    if document_candidates:log(f"  PDF CANDIDATOS: {len(document_candidates)}")
    for candidate in document_candidates:
        if candidate.url in seen_urls:continue
        if budget_tracker and budget_tracker.pdfs_analyzed>=budget_tracker.budget.max_pdfs_analyzed_per_product:break
        if budget_tracker and not budget_tracker.can_accept_source():break
        seen_urls.add(candidate.url)
        if budget_tracker:budget_tracker.pdfs_analyzed+=1
        try:
            doc_rec=process_pdf_document(identity,candidate.url,target_semantics=target_semantics,trace=trace)
            if not _log_source_decision(doc_rec,log):raise ValueError("SOURCE_VALIDATION_REJECTED: NO_POLICY_APPROVED_EVIDENCE")
            if budget_tracker and not budget_tracker.accept_source():break
            accepted.append(doc_rec);log(f"  PDF VALIDADO: {candidate.url}")
            if len(accepted)>=max(0,int(accept_limit)):break
        except Exception as exc:
            errors.append(f"document:{candidate.url}: {type(exc).__name__}: {exc}");_log_source_rejection(exc,log)
    for line in format_trace_lines(trace):log(line)
    return accepted


def scrape_item(
    item: BatchItem,
    out_dir: str,
    template_plan: dict | None = None,
    log=lambda m: None,
    source_strategy: SourceStrategy | None = None,
) -> ProductRecord | None:
    strategy = (source_strategy or SourceStrategy()).normalized()
    pipe=ProductPipeline()
    media_slots=int((template_plan or {}).get("media_slots",0) or 0);target_semantics=list((template_plan or {}).get("scrape_semantics") or []);include_images=bool(media_slots) and strategy.web;include_pdfs=strategy.pdf and bool((template_plan or {}).get("summary",{}).get("scrape_targets",1))
    classification=classify_product(item.identity,required_fields=target_semantics) if target_semantics else None
    orchestrator=ProductEvidenceOrchestrator(item.identity,target_semantics,category=classification.category if classification else None,budget=RUNTIME_RESOLUTION_BUDGET,source_strategy=strategy) if target_semantics else None
    budget_tracker=orchestrator.budget_tracker if orchestrator else SearchBudgetTracker(RUNTIME_RESOLUTION_BUDGET)
    manual_urls=list(dict.fromkeys([u for u in ((item.source_urls or [])+([item.source_url] if item.source_url else [])) if u]));eligible_manual_urls=[u for u in manual_urls if strategy.web or (strategy.pdf and _looks_like_pdf_url(u))]
    manual_candidates=[type("Candidate",(),{"url":u,"likely_official":False,"score":2.0,"ai_assisted":False,"manual_source":True})() for u in eligible_manual_urls]
    smart_snapshot=orchestrator.plan_next() if orchestrator else None
    first_external=next((intent for intent in (smart_snapshot.next_intents if smart_snapshot else ()) if intent.engine!="EXISTING"),None)
    prefer_pdf_first=bool(first_external and first_external.engine=="PDF")
    should_search_initial_web=bool(strategy.web and not eligible_manual_urls and not prefer_pdf_first)
    free_candidates=search_web(item.identity,limit=RUNTIME_RESOLUTION_BUDGET.max_candidates_per_query,budget_tracker=budget_tracker,query_quota=SEARCH_STAGE_QUERY_QUOTAS["INITIAL"]) if should_search_initial_web else []
    if orchestrator and should_search_initial_web:log(f"  SMART QUERY: used={budget_tracker.queries_used} limit={RUNTIME_RESOLUTION_BUDGET.max_search_queries_per_product} engine={first_external.engine if first_external else 'WEB_STRUCTURED'}")
    known_manual=set(eligible_manual_urls);candidates=manual_candidates+_prioritize([c for c in free_candidates if getattr(c,"url",None) not in known_manual])
    if manual_urls:log(f"  fuentes manuales: {len(manual_urls)}; elegibles por estrategia={len(eligible_manual_urls)}; se validan antes de aceptar evidencia")
    if orchestrator:
        log(f"  SMART IDENTITY: {item.identity.brand or '-'} / {item.identity.model or item.identity.product_name or item.identity.mpn or '-'}")
        log(f"  SMART PLAN: requeridos={len(target_semantics)} pendientes={len(smart_snapshot.missing_fields)} next={(first_external.engine if first_external else 'STOP')} category={orchestrator.category}")
    errors=[];accepted=[];queue=list(candidates);seen_urls={getattr(c,"url","") for c in queue};manufacturer_followup_done=False;cursor=0

    while cursor<len(queue) and len(accepted)<MAX_VALIDATED_SOURCES_PER_PRODUCT and budget_tracker.can_accept_source():
        candidate=queue[cursor];cursor+=1
        if budget_tracker.pages_fetched>=RUNTIME_RESOLUTION_BUDGET.max_pages_fetched_per_product:
            log("  STOP: PAGE_FETCH_BUDGET agotado");break
        try:
            is_pdf=bool(strategy.pdf and _looks_like_pdf_url(candidate.url))
            if orchestrator:
                fields_text=",".join(smart_snapshot.missing_fields or tuple(target_semantics))
                source_engine="PDF" if is_pdf else "WEB_STRUCTURED"
                source_kind="MANUAL_PDF" if is_pdf and getattr(candidate,"manual_source",False) else "CANDIDATE"
                log(f"  SMART SOURCE: {source_engine} kind={source_kind} fields={fields_text}")
            log(f"  probando: {candidate.url}")
            if is_pdf:
                if budget_tracker.pdfs_analyzed>=RUNTIME_RESOLUTION_BUDGET.max_pdfs_analyzed_per_product:continue
                budget_tracker.pdfs_analyzed+=1;rec=process_pdf_document(item.identity,candidate.url,target_semantics=target_semantics)
            else:
                if not strategy.web:continue
                budget_tracker.pages_fetched+=1;official_domain=_candidate_official_domain(candidate,item.identity,accepted)
                rec=pipe.process_url(item.identity,candidate.url,official_domain=official_domain,include_pdfs=include_pdfs,include_images=include_images,browser_fallback=True,target_semantics=target_semantics,media_slots=media_slots)
            if not _log_source_decision(rec,log):raise ValueError("SOURCE_VALIDATION_REJECTED: NO_POLICY_APPROVED_EVIDENCE")
            if rec.identity.identifiers_conflicting:raise ValueError("SOURCE_VALIDATION_REJECTED: IDENTITY_CONFLICT legacy_identifiers")
            if accepted and not _cross_source_consistent(accepted[0],rec,candidate.url):raise ValueError("SOURCE_VALIDATION_REJECTED: CROSS_SOURCE_PRODUCT_CONFLICT")
            if orchestrator:
                accepted.append(rec)
                smart_snapshot=orchestrator.observe_record(rec,engine="PDF" if is_pdf else "WEB_STRUCTURED",source_url=candidate.url,source_kind=(rec.fetch or {}).get("source_class"))
                log(f"  SMART FIELDS: verificados={len(smart_snapshot.resolved_fields)} faltantes={len(smart_snapshot.missing_fields)} conflictos={len(smart_snapshot.conflicted_fields)}")
            else:
                if not budget_tracker.accept_source():break
                accepted.append(rec)
            log(f"  fuente validada: {(rec.fetch or {}).get('source_class','?')} / {rec.identity.match_level}")

            if orchestrator and smart_snapshot.early_stop:
                log(f"  SMART FINAL: STOP reason={smart_snapshot.stop_reason}")
                break

            needs_identity_refinement=not orchestrator or any(intent.engine=="IDENTITY" for intent in smart_snapshot.next_intents)
            if strategy.web and not manufacturer_followup_done and budget_tracker.remaining_queries()>0 and needs_identity_refinement:
                enriched=_enriched_identity(item,rec)
                if enriched:
                    followups=search_web(enriched,limit=RUNTIME_RESOLUTION_BUDGET.max_candidates_per_query,budget_tracker=budget_tracker,query_quota=SEARCH_STAGE_QUERY_QUOTAS["IDENTITY_REFINEMENT"]);fresh=[c for c in followups if c.url not in seen_urls]
                    if orchestrator:log(f"  SMART QUERY: used={budget_tracker.queries_used} limit={RUNTIME_RESOLUTION_BUDGET.max_search_queries_per_product} engine=IDENTITY")
                    for c in fresh:seen_urls.add(c.url)
                    queue[cursor:cursor]=_prioritize(fresh);log(f"  búsqueda fabricante reforzada: {len(fresh)} candidatos nuevos")
                manufacturer_followup_done=True

            has_manufacturer=any((r.fetch or {}).get("source_class")=="manufacturer" for r in accepted)
            if has_manufacturer:
                _merged,partial_resolution=_resolution_for(accepted,template_plan);remaining=list(partial_resolution.get("research_terms") or []);log(f"  cobertura actual: {len(target_semantics)-len(remaining)}/{len(target_semantics)} semánticas; pendientes={len(remaining)}")
                if not orchestrator and _coverage_sufficient(partial_resolution,has_manufacturer=True):log("  cobertura suficiente con fabricante validado; se detiene PASS 1");break
        except Exception as exc:
            errors.append(f"{candidate.url}: {type(exc).__name__}: {exc}");_log_source_rejection(exc,log)

    if not accepted and strategy.pdf:
        wants_pdf=not orchestrator or any(intent.engine=="PDF" for intent in (smart_snapshot.next_intents if smart_snapshot else ()))
        if wants_pdf and budget_tracker.can_accept_source():
            if orchestrator:log(f"  SMART SOURCE: PDF kind=OFFICIAL_PDF fields={','.join(smart_snapshot.missing_fields or tuple(target_semantics))}")
            direct_records=_ingest_direct_documents(item.identity,target_semantics=target_semantics,seen_urls=seen_urls,errors=errors,log=log,budget_tracker=budget_tracker,accept_limit=min(3,RUNTIME_RESOLUTION_BUDGET.max_sources_accepted_per_product))
            accepted.extend(direct_records)
            if orchestrator:
                for doc_rec in direct_records:
                    source_url=(doc_rec.fetch or {}).get("final_url") or next(iter(doc_rec.sources or []),"PDF_DISCOVERY")
                    smart_snapshot=orchestrator.observe_record(doc_rec,engine="PDF",source_url=source_url,source_kind=(doc_rec.fetch or {}).get("source_class"),count_source=False)
                    log(f"  SMART FIELDS: verificados={len(smart_snapshot.resolved_fields)} faltantes={len(smart_snapshot.missing_fields)} conflictos={len(smart_snapshot.conflicted_fields)}")
    if not accepted:
        if orchestrator:
            orchestrator.observe_source_outcome(first_external,"NO_RESULT",engine=first_external.engine if first_external else "UNKNOWN",reason="NO_VALIDATED_SOURCE")
        log("  SIN FUENTE VALIDADA: "+(errors[-1] if errors else "no hubo candidatos"));return None

    rec=_merge_valid_records(accepted);resolution=analyze_resolution(rec,template_plan)
    smart_snapshot=orchestrator.plan_next() if orchestrator else None
    gap_terms=list(smart_snapshot.missing_fields if smart_snapshot else (resolution.get("research_terms") or []))
    wants_more=not smart_snapshot or not smart_snapshot.early_stop
    wants_pdf=not orchestrator or any(intent.engine=="PDF" for intent in (smart_snapshot.next_intents if smart_snapshot else ()))
    if gap_terms and strategy.pdf and wants_more and wants_pdf and budget_tracker.can_accept_source():
        remaining_sources=RUNTIME_RESOLUTION_BUDGET.max_sources_accepted_per_product-len(accepted)
        if orchestrator:log(f"  SMART NEXT_SOURCE: PDF kind=OFFICIAL_PDF fields={','.join(gap_terms)}")
        document_extra=_ingest_direct_documents(rec.identity,target_semantics=gap_terms,seen_urls=seen_urls,errors=errors,log=log,budget_tracker=budget_tracker,accept_limit=max(0,min(3,remaining_sources)))
        if document_extra:
            accepted.extend(document_extra)
            if orchestrator:
                for doc_rec in document_extra:
                    source_url=(doc_rec.fetch or {}).get("final_url") or next(iter(doc_rec.sources or []),"PDF_DISCOVERY")
                    smart_snapshot=orchestrator.observe_record(doc_rec,engine="PDF",source_url=source_url,source_kind=(doc_rec.fetch or {}).get("source_class"),count_source=False)
                    log(f"  SMART FIELDS: verificados={len(smart_snapshot.resolved_fields)} faltantes={len(smart_snapshot.missing_fields)} conflictos={len(smart_snapshot.conflicted_fields)}")
            rec=_merge_valid_records(accepted);resolution=analyze_resolution(rec,template_plan);gap_terms=list(smart_snapshot.missing_fields if smart_snapshot else (resolution.get("research_terms") or []));log(f"  cobertura tras documentos: pendientes={len(gap_terms)}")

    smart_snapshot=orchestrator.plan_next() if orchestrator else None
    gap_terms=list(smart_snapshot.missing_fields if smart_snapshot else (resolution.get("research_terms") or []))
    wants_web=not orchestrator or any(intent.engine in {"IDENTITY","WEB_STRUCTURED","WEB_FALLBACK"} for intent in (smart_snapshot.next_intents if smart_snapshot else ()))
    if strategy.web and gap_terms and wants_web and (not smart_snapshot or not smart_snapshot.early_stop) and budget_tracker.remaining_queries()>0 and budget_tracker.can_accept_source():
        mode="conflictos/huecos" if resolution.get("blocked") else "huecos";log(f"  segunda pasada por {mode}: {len(gap_terms)} campos/grupos pendientes");current_sources=set(rec.sources or []);total_extra=0;max_extra=max(0,RUNTIME_RESOLUTION_BUDGET.max_sources_accepted_per_product-len(accepted))
        next_web_intent=next((intent for intent in (smart_snapshot.next_intents if smart_snapshot else ()) if intent.engine in {"IDENTITY","WEB_STRUCTURED","WEB_FALLBACK"}),None)
        if orchestrator and next_web_intent:log(f"  SMART NEXT_SOURCE: {next_web_intent.engine} kind={next_web_intent.source_kind} fields={','.join(gap_terms)}")
        for start in range(0,len(gap_terms),4):
            if total_extra>=max_extra or budget_tracker.remaining_queries()<=0 or not budget_tracker.can_accept_source():break
            chunk=gap_terms[start:start+4];log(f"    buscando específicamente: {', '.join(chunk)}")
            raw_chunk_candidates=search_web_for_fields(rec.identity,chunk,limit=RUNTIME_RESOLUTION_BUDGET.max_candidates_per_query,budget_tracker=budget_tracker,query_quota=min(SEARCH_STAGE_QUERY_QUOTAS["MISSING_FIELDS"],budget_tracker.remaining_queries()),source_kind=next_web_intent.source_kind if next_web_intent else None,category=orchestrator.category if orchestrator else None)
            if orchestrator:log(f"  SMART QUERY: used={budget_tracker.queries_used} limit={RUNTIME_RESOLUTION_BUDGET.max_search_queries_per_product} engine={next_web_intent.engine if next_web_intent else 'WEB_STRUCTURED'}")
            chunk_candidates=_prioritize(raw_chunk_candidates);chunk_extra=[]
            for candidate in chunk_candidates:
                if total_extra>=max_extra or not budget_tracker.can_accept_source():break
                if candidate.url in seen_urls or candidate.url in current_sources:continue
                if budget_tracker.pages_fetched>=RUNTIME_RESOLUTION_BUDGET.max_pages_fetched_per_product:break
                seen_urls.add(candidate.url)
                try:
                    budget_tracker.pages_fetched+=1;official_domain=_candidate_official_domain(candidate,rec.identity,accepted)
                    gap_rec=pipe.process_url(item.identity,candidate.url,official_domain=official_domain,include_pdfs=include_pdfs,include_images=include_images,browser_fallback=True,target_semantics=chunk,media_slots=media_slots)
                    if not _log_source_decision(gap_rec,log,prefix="    "):raise ValueError("SOURCE_VALIDATION_REJECTED: NO_POLICY_APPROVED_EVIDENCE")
                    if gap_rec.identity.identifiers_conflicting:raise ValueError("SOURCE_VALIDATION_REJECTED: IDENTITY_CONFLICT legacy_identifiers")
                    if accepted and not _cross_source_consistent(accepted[0],gap_rec,candidate.url):raise ValueError("SOURCE_VALIDATION_REJECTED: CROSS_SOURCE_PRODUCT_CONFLICT")
                    if orchestrator:
                        chunk_extra.append(gap_rec);accepted.append(gap_rec);total_extra+=1
                        smart_snapshot=orchestrator.observe_record(gap_rec,engine="WEB_STRUCTURED",source_url=candidate.url,source_kind=(gap_rec.fetch or {}).get("source_class"))
                        log(f"  SMART FIELDS: verificados={len(smart_snapshot.resolved_fields)} faltantes={len(smart_snapshot.missing_fields)} conflictos={len(smart_snapshot.conflicted_fields)}")
                    else:
                        if not budget_tracker.accept_source():break
                        chunk_extra.append(gap_rec);accepted.append(gap_rec);total_extra+=1
                    log(f"    gap fuente validada: {(gap_rec.fetch or {}).get('source_class','?')} / {gap_rec.identity.match_level}")
                    if orchestrator and smart_snapshot.early_stop:break
                    if len(chunk_extra)>=2:break
                except Exception as exc:
                    errors.append(f"gap:{candidate.url}: {type(exc).__name__}: {exc}");_log_source_rejection(exc,log,prefix="    ")
            if chunk_extra:
                rec=_merge_valid_records(accepted);resolution=analyze_resolution(rec,template_plan);gap_terms=list(smart_snapshot.missing_fields if orchestrator else (resolution.get("research_terms") or []))
                if not gap_terms and not resolution.get("blocked"):log("  PASS 3 completó todos los huecos resolubles");break
                if orchestrator and smart_snapshot.early_stop:break

    rec.evidence_graph=dict(rec.evidence_graph or {});rec.evidence_graph["resolution_audit"]=resolution;rec.evidence_graph["resolution_budget"]={"queries_used":budget_tracker.queries_used,"candidates_admitted":budget_tracker.candidates_admitted,"pages_fetched":budget_tracker.pages_fetched,"pdfs_analyzed":budget_tracker.pdfs_analyzed,"sources_accepted":budget_tracker.sources_accepted,"limits":RUNTIME_RESOLUTION_BUDGET.__dict__}
    if orchestrator:
        smart_snapshot=orchestrator.plan_next();rec.evidence_graph["smart_orchestrator"]=orchestrator.audit();rec.missing_fields=list(smart_snapshot.missing_fields)
        log(f"  SMART FINAL: verified={len(smart_snapshot.resolved_fields)}/{len(smart_snapshot.required_fields)} missing={len(smart_snapshot.missing_fields)} conflicts={len(smart_snapshot.conflicted_fields)} stop={smart_snapshot.stop_reason or 'NO'}")
    else:
        rec.missing_fields=[row["semantic"] for row in resolution.get("fields",[]) if row.get("status")=="INSUFFICIENT_EVIDENCE"]
    for issue in resolution.get("cross_field_issues",[]):rec.warnings.append(f"cross_field:{issue.get('code')}")
    if resolution.get("blocked"):rec.warnings.append("final_material_conflict_after_targeted_research")
    Path(out_dir).mkdir(parents=True,exist_ok=True);stem=re.sub(r"[^A-Za-z0-9._-]+","_",item.identity.mpn or item.identity.ean or item.identity.model or f"row_{item.row}");(Path(out_dir)/f"{stem}.json").write_text(rec.model_dump_json(indent=2),encoding="utf-8")
    return rec


def run_batch(
    template: str,
    output_dir: str,
    overwrite: bool = False,
    log=lambda m: None,
    manual_part_numbers: list[str] | None = None,
    manual_identities: list[ProductIdentity] | None = None,
    manual_source_urls: list[list[str]] | None = None,
    source_strategy: SourceStrategy | None = None,
) -> dict:
    strategy = (source_strategy or SourceStrategy()).normalized()
    out=Path(output_dir);out.mkdir(parents=True,exist_ok=True);template_plan=analyze_template_contract(template);(out/"template_contract.json").write_text(json.dumps(template_plan,ensure_ascii=False,indent=2),encoding="utf-8");ps=template_plan["summary"]
    log(f"Contrato Excel: {ps['fields_total']} campos | {ps['scrape_targets']} datos de producto | {ps['media_slots']} imágenes | {ps['seller_inputs']} datos del vendedor | {ps['marketplace_inputs']} datos marketplace")
    log("Fuentes ejecución: "+f"WEB={'ON' if strategy.web else 'OFF'} | PDF={'ON' if strategy.pdf else 'OFF'} | OCR={'ON' if strategy.ocr else 'OFF'} | MISTRAL={'ON' if strategy.mistral else 'OFF'}")
    manual_mode=bool(manual_identities or manual_part_numbers)
    if manual_identities:items=manual_identity_items(template,manual_identities,manual_source_urls)
    elif manual_part_numbers:items=manual_items(template,manual_part_numbers)
    else:items=detect_items(template)
    log(f"Productos a procesar: {len(items)}"+(" (entradas manuales: MPN/EAN/UPC/GTIN/nombre)" if manual_mode else " (detectados en Excel)"))
    records=[];row_assignments={};failures=[]
    for index,item in enumerate(items,1):
        label=item.identity.mpn or item.identity.ean or item.identity.model or item.identity.product_name;log(f"[{index}/{len(items)}] {label}");rec=scrape_item(item,str(out/"json"),template_plan=template_plan,log=log,source_strategy=strategy)
        if rec:records.append(rec);row_assignments[(item.sheet,item.row)]=rec
        else:failures.append({"part_number":label,"sheet":item.sheet,"row":item.row})
    output_xlsx=str(out/(Path(template).stem+"_completado.xlsx"));trace=str(out/"trazabilidad.json");report=fill_excel_v8(template,output_xlsx,[],overwrite=overwrite,trace_path=trace,row_assignments=row_assignments)
    resolution_summary={str(item.identity.mpn or item.identity.ean or item.identity.model or item.identity.product_name):(row_assignments[(item.sheet,item.row)].evidence_graph or {}).get("resolution_audit",{}) for item in items if (item.sheet,item.row) in row_assignments};resolution_file=out/"resolucion_campos.json";resolution_file.write_text(json.dumps(resolution_summary,ensure_ascii=False,indent=2),encoding="utf-8")
    source_validation_summary={"validated_sources":sum(int((rec.fetch or {}).get("validated_sources",1) or 1) for rec in records),"manufacturer_sources":sum(int((rec.fetch or {}).get("manufacturer_sources",1 if (rec.fetch or {}).get("source_class")=="manufacturer" else 0) or 0) for rec in records),"material_evidence":sum(len(rec.evidence or []) for rec in records),"final_specifications":sum(len(rec.specifications or {}) for rec in records)}
    summary={"mode":"manual_product_identity" if manual_mode else "excel_detected","source_strategy":strategy.as_options(),"template_contract":template_plan["summary"],"template_contract_file":str(out/"template_contract.json"),"products_detected":len(items),"products_scraped":len(records),"products_failed":len(failures),"failures":failures,"source_validation":source_validation_summary,"output_excel":output_xlsx,"trace":trace,"resolution":str(resolution_file),"mapping":report.get("summary",{})}
    (out/"resumen.json").write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding="utf-8");return summary
