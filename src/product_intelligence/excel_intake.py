from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ProductIdentity
from .normalize import canonical_key, key_norm
from .semantic_guard import is_placeholder


_TEMPLATE_IDENTITY_EXAMPLES = {
    "1234567890",
    "99999999",
    "999999999",
    "abc-1000-202",
    "abc1000202",
}

_IDENTITY_ALIASES = {
    "part_number": {
        "mpn", "pn", "part number", "partnumber", "part no", "part no.",
        "manufacturer part number", "manufacturer partnumber", "manufacturer pn",
        "codigo fabricante", "código fabricante", "numero de parte", "número de parte",
        "modelo fabricante", "manufacturer model",
    },
    "gtin": {
        "ean", "ean13", "ean 13", "upc", "upc a", "gtin", "gtin14", "gtin 14",
        "ean upc", "ean/upc", "barcode", "codigo de barras", "código de barras",
        "global trade item number",
    },
    "sku": {
        "sku", "seller sku", "sku seller", "sku_seller", "seller_sku",
        "merchant sku", "merchant_sku", "sku vendedor", "sku del vendedor",
    },
    "model": {
        "model", "modelo", "model number", "model no", "model no.",
        "numero de modelo", "número de modelo",
    },
    "brand": {"brand", "marca", "manufacturer", "fabricante"},
    "product_name": {
        "product name", "nombre producto", "nombre del producto", "name", "nombre",
        "title", "titulo", "título",
    },
    "source_url": {
        "source url", "url fuente", "pagina oficial", "página oficial", "official url",
        "product url", "url producto",
    },
}


def normalize_header(label: Any) -> str:
    text = re.sub(r"#\s*[A-Za-z]*\d+", "", str(label or ""))
    text = key_norm(text)
    return re.sub(r"\s+", " ", text).strip()


def _compact_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_header(value))


_ALIAS_COMPACT = {
    kind: {_compact_header(alias) for alias in aliases}
    for kind, aliases in _IDENTITY_ALIASES.items()
}


def identity_header_kind(label: Any) -> str | None:
    normalized = normalize_header(label)
    compact = _compact_header(normalized)
    if not normalized:
        return None
    for kind, aliases in _IDENTITY_ALIASES.items():
        if normalized in {normalize_header(x) for x in aliases} or compact in _ALIAS_COMPACT[kind]:
            return kind
    canonical = canonical_key(normalized)
    mapping = {
        "mpn": "part_number",
        "ean": "gtin",
        "upc": "gtin",
        "gtin": "gtin",
        "model": "model",
        "brand": "brand",
    }
    return mapping.get(canonical)


def _clean_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text or None


def _identity_placeholder(value: Any) -> bool:
    text = _clean_value(value)
    if not text:
        return True
    compact = key_norm(text).replace(" ", "")
    return compact in _TEMPLATE_IDENTITY_EXAMPLES or is_placeholder(text)


def looks_like_part_number(value: Any) -> bool:
    text = str(value or "").strip()
    if not text or " " in text or len(text) < 4 or len(text) > 80:
        return False
    if not re.fullmatch(r"[A-Za-z0-9._/-]+", text):
        return False
    if text.isdigit():
        return False
    return bool(re.search(r"[A-Za-z]", text) and re.search(r"\d", text))


def _normalize_gtin(value: Any) -> str | None:
    text = _clean_value(value)
    if not text or _identity_placeholder(text):
        return None
    compact = re.sub(r"[\s-]+", "", text)
    if compact.isdigit() and len(compact) in {8, 12, 13, 14}:
        return compact
    return text


def _value_looks_valid(kind: str, value: Any) -> bool:
    text = _clean_value(value)
    if not text or _identity_placeholder(text):
        return False
    if kind == "gtin":
        compact = re.sub(r"[\s-]+", "", text)
        return compact.isdigit() and len(compact) in {8, 12, 13, 14}
    if kind in {"part_number", "sku"}:
        return looks_like_part_number(text) or (len(text) >= 3 and len(text) <= 80 and " " not in text)
    if kind == "model":
        return len(text) <= 120
    if kind == "product_name":
        return 2 <= len(text) <= 250
    return bool(text)


@dataclass
class IntakeProduct:
    sheet: str
    row: int
    identity: ProductIdentity
    source_url: str | None = None
    source_urls: list[str] = field(default_factory=list)
    audit: dict[str, Any] = field(default_factory=dict)


@dataclass
class SheetAudit:
    sheet: str
    score: float = 0.0
    accepted: bool = False
    header_row: int | None = None
    raw_headers: dict[int, str] = field(default_factory=dict)
    normalized_headers: dict[int, str] = field(default_factory=dict)
    identity_columns: dict[str, list[int]] = field(default_factory=dict)
    rejection_reason: str | None = None
    rows: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_detected": self.sheet,
            "sheet_score": self.score,
            "sheet_accepted": self.accepted,
            "header_row_detected": self.header_row,
            "raw_headers": self.raw_headers,
            "normalized_headers": self.normalized_headers,
            "identity_columns": self.identity_columns,
            "sheet_rejection_reason": self.rejection_reason,
            "rows": self.rows,
        }


@dataclass
class WorkbookIntakeResult:
    products: list[IntakeProduct] = field(default_factory=list)
    sheets: list[SheetAudit] = field(default_factory=list)

    def audit_dict(self) -> dict[str, Any]:
        return {"sheets": [sheet.to_dict() for sheet in self.sheets]}


def _candidate_header(ws, row: int) -> tuple[float, dict[int, str], dict[int, str], dict[str, list[int]], int]:
    raw_headers: dict[int, str] = {}
    normalized_headers: dict[int, str] = {}
    columns: dict[str, list[int]] = {}
    semantic_fields = 0

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value in (None, ""):
            continue
        raw = str(value).strip()
        normalized = normalize_header(raw)
        raw_headers[col] = raw
        normalized_headers[col] = normalized
        kind = identity_header_kind(raw)
        if kind:
            columns.setdefault(kind, []).append(col)
        if canonical_key(normalized):
            semantic_fields += 1

    identity_kinds = {k for k in columns if k not in {"brand", "source_url"}}
    product_rows = 0
    evidence_hits = 0
    if identity_kinds:
        end = min(ws.max_row, row + 30)
        for rr in range(row + 1, end + 1):
            row_hit = False
            for kind in identity_kinds:
                for col in columns.get(kind, []):
                    if _value_looks_valid(kind, ws.cell(rr, col).value):
                        evidence_hits += 1
                        row_hit = True
                        break
                if row_hit:
                    break
            if row_hit:
                product_rows += 1

    header_score = len(identity_kinds) * 8 + semantic_fields * 1.5 + min(len(raw_headers), 12) * 0.15
    score = header_score + min(product_rows, 10) * 4 + min(evidence_hits, 15) * 0.5
    return score, raw_headers, normalized_headers, columns, product_rows


def analyze_sheet(ws) -> SheetAudit:
    best: tuple[float, int, dict[int, str], dict[int, str], dict[str, list[int]], int] | None = None
    scan_end = min(ws.max_row, 30)
    for row in range(1, scan_end + 1):
        score, raw, normalized, columns, product_rows = _candidate_header(ws, row)
        if best is None or score > best[0]:
            best = (score, row, raw, normalized, columns, product_rows)

    if best is None or not best[2]:
        return SheetAudit(sheet=ws.title, rejection_reason="NO_HEADER")

    score, header_row, raw, normalized, columns, product_rows = best
    identity_columns = {k: v for k, v in columns.items() if k not in {"brand", "source_url"}}
    if not identity_columns:
        return SheetAudit(
            sheet=ws.title,
            score=score,
            header_row=header_row,
            raw_headers=raw,
            normalized_headers=normalized,
            identity_columns=columns,
            rejection_reason="NO_IDENTITY_COLUMN",
        )
    if product_rows == 0:
        return SheetAudit(
            sheet=ws.title,
            score=score,
            header_row=header_row,
            raw_headers=raw,
            normalized_headers=normalized,
            identity_columns=columns,
            rejection_reason="NO_PRODUCT_ROWS",
        )
    return SheetAudit(
        sheet=ws.title,
        score=score,
        accepted=True,
        header_row=header_row,
        raw_headers=raw,
        normalized_headers=normalized,
        identity_columns=columns,
    )


def _first_value(ws, row: int, columns: list[int]) -> tuple[int | None, str | None]:
    for col in columns:
        value = _clean_value(ws.cell(row, col).value)
        if value and not _identity_placeholder(value):
            return col, value
    return None, None


def _resolve_row(ws, audit: SheetAudit, row: int) -> IntakeProduct | None:
    columns = audit.identity_columns
    values: dict[str, str] = {}
    raw_identity: dict[str, Any] = {}
    selected_type = None
    selected_value = None

    part_col, part = _first_value(ws, row, columns.get("part_number", []))
    gtin_col, gtin = _first_value(ws, row, columns.get("gtin", []))
    model_col, model = _first_value(ws, row, columns.get("model", []))
    sku_col, sku = _first_value(ws, row, columns.get("sku", []))
    name_col, product_name = _first_value(ws, row, columns.get("product_name", []))
    brand_col, brand = _first_value(ws, row, columns.get("brand", []))
    _source_col, source_url = _first_value(ws, row, columns.get("source_url", []))

    for kind, col, value in [
        ("part_number", part_col, part), ("gtin", gtin_col, gtin), ("model", model_col, model),
        ("sku", sku_col, sku), ("product_name", name_col, product_name), ("brand", brand_col, brand),
    ]:
        if value is not None:
            raw_identity[kind] = {"column": col, "value": value}

    if part:
        values["mpn"] = part
        selected_type, selected_value = "PART_NUMBER", part
    elif gtin:
        normalized_gtin = _normalize_gtin(gtin)
        if normalized_gtin:
            values["gtin"] = normalized_gtin
            selected_type, selected_value = "GTIN", normalized_gtin
    elif model:
        values["model"] = model
        if looks_like_part_number(model):
            values["mpn"] = model
            selected_type, selected_value = "PART_NUMBER_FROM_MODEL", model
        else:
            selected_type, selected_value = "MODEL", model
    elif sku:
        values["sku"] = sku
        selected_type, selected_value = "SKU", sku
    elif product_name:
        values["product_name"] = product_name
        selected_type, selected_value = "PRODUCT_NAME", product_name

    if brand:
        values["brand"] = brand
    if model and "model" not in values:
        values["model"] = model
    if product_name and "product_name" not in values:
        values["product_name"] = product_name
    if sku and "sku" not in values:
        values["sku"] = sku

    row_audit = {
        "product_row": row,
        "part_number_column": part_col,
        "gtin_column": gtin_col,
        "sku_column": sku_col,
        "part_number_raw": part,
        "gtin_raw": gtin,
        "sku_raw": sku,
        "identity_selected": selected_value,
        "identity_type": selected_type,
        "row_accepted": bool(selected_value),
        "rejection_reason": None if selected_value else "NO_IDENTITY_VALUE",
        "raw_identity": raw_identity,
    }
    audit.rows.append(row_audit)

    if not selected_value:
        return None
    identity = ProductIdentity(**values)
    return IntakeProduct(
        sheet=ws.title,
        row=row,
        identity=identity,
        source_url=source_url,
        source_urls=[source_url] if source_url else [],
        audit=row_audit,
    )


def analyze_workbook_intake(template: str) -> WorkbookIntakeResult:
    path = Path(template)
    if not path.exists():
        raise FileNotFoundError(template)
    wb = load_workbook(template, data_only=False, read_only=False)
    result = WorkbookIntakeResult()
    for ws in wb.worksheets:
        audit = analyze_sheet(ws)
        result.sheets.append(audit)
        if not audit.accepted or audit.header_row is None:
            continue
        for row in range(audit.header_row + 1, ws.max_row + 1):
            product = _resolve_row(ws, audit, row)
            if product is not None:
                result.products.append(product)
    return result
