from __future__ import annotations

import json, re
from pathlib import Path
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from .normalize import ALIASES, key_norm, canonical_key
from .models import ProductRecord
from .template_intelligence import classify_field
from .marketplace_mapper import derive_template_value, media_rank
from .semantic_guard import infer_contract, validate_value

IDENTITY_ALIASES={
    "mpn":["mpn","part number","part no","pn","codigo fabricante","código fabricante"],
    "ean":["ean","ean13","ean-13","codigo barras","código de barras","barcode"],
    "upc":["upc"], "gtin":["gtin"], "sku":["sku"], "model":["model","modelo"], "brand":["brand","marca"]
}


def _external_id(header: str) -> str | None:
    m = re.search(r"#\s*([A-Za-z]*\d+|\d+)", str(header))
    return m.group(1) if m else None


def map_header(header: str, description: str | None = None) -> tuple[str|None,float,str]:
    raw = re.sub(r"#\s*[A-Za-z]*\d+", "", str(header)).strip()
    h=key_norm(raw)
    cls,_,class_conf=classify_field(str(header),description)
    if cls=="IMAGE": return "__image__",class_conf,cls
    if cls=="SELLER_DATA": return None,class_conf,cls
    candidates=[]
    for k,als in {**IDENTITY_ALIASES, **ALIASES}.items():
        candidates += [(key_norm(a),k) for a in [k,*als]]
    exact=next((k for a,k in candidates if a==h),None)
    if exact: return exact,1.0,cls

    # V6: fuzzy mapping is intentionally conservative. Long marketplace descriptions are
    # used to infer a contract, NOT concatenated into the alias query, because that was
    # the source of cross-attribute mappings (cable length -> product length, etc.).
    best=process.extractOne(h,[a for a,_ in candidates],scorer=fuzz.ratio)
    if not best or best[1] < 88: return None,best[1]/100 if best else 0,cls
    alias=best[0]; key=next(k for a,k in candidates if a==alias)
    return key,best[1]/100,cls


def _record_value(rec: ProductRecord, key: str):
    if key in IDENTITY_ALIASES:
        return getattr(rec.identity,key,None), None, rec.identity.confidence, None, None
    d=rec.specifications.get(key)
    if not d: return None,None,0,None,None
    return d.get("value"), d.get("source"), d.get("confidence",0), key, d.get("raw_value")


def _detect_header_and_description(ws):
    best=(1,-1.0)
    for r in range(1,min(ws.max_row,25)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        nonempty=[str(v) for v in vals if v not in (None,"")]
        if not nonempty: continue
        score=0.0
        for v in nonempty:
            cls,_,_=classify_field(v)
            if canonical_key(re.sub(r"#\s*[A-Za-z]*\d+","",v).strip()): score+=3
            if re.search(r"#\s*[A-Za-z]*\d+",v): score+=2
            if cls in {"IMAGE","SELLER_DATA","DERIVABLE"}: score+=1
            if len(v)<80: score+=0.2
        if score>best[1]: best=(r,score)
    header=best[0]
    desc=None; desc_score=-1
    for r in range(max(1,header-3),header):
        vals=[str(ws.cell(r,c).value) for c in range(1,ws.max_column+1) if ws.cell(r,c).value not in (None,"")]
        avg=sum(map(len,vals))/len(vals) if vals else 0
        if avg>desc_score: desc,desc_score=r,avg
    return header,desc


def _candidate_evidence(rec: ProductRecord, key: str):
    # Return strongest exact/high evidence for contextual validation.
    evs=[e for e in rec.evidence if canonical_key(e.attribute)==key]
    evs.sort(key=lambda e:e.confidence, reverse=True)
    return evs[0] if evs else None



def _strip_field_id(label: str) -> str:
    return re.sub(r"#\s*[A-Za-z]*\d+", "", str(label)).strip()


def _build_option_index(wb) -> dict[str, list]:
    """Discover controlled vocabularies from any sheet that looks like an options table."""
    index: dict[str, list] = {}
    for ws in wb.worksheets:
        # Look at first five rows for a header row with several compact labels.
        for r in range(1, min(ws.max_row, 5) + 1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            compact=[v for v in vals if v not in (None,"") and len(str(v)) < 80]
            if len(compact) < 2:
                continue
            # An options table normally has many values below each header.
            populated_cols=0
            for c,v in enumerate(vals,1):
                if v in (None,""): continue
                below=sum(1 for rr in range(r+1,min(ws.max_row,r+60)+1) if ws.cell(rr,c).value not in (None,""))
                if below>=2: populated_cols+=1
            if populated_cols<2:
                continue
            for c,h in enumerate(vals,1):
                if h in (None,""): continue
                options=[]
                for rr in range(r+1,ws.max_row+1):
                    v=ws.cell(rr,c).value
                    if v not in (None,""):
                        options.append(v)
                if options:
                    index[key_norm(str(h))]=options
            break
    return index


def _find_options_for_field(option_index: dict[str,list], header: str) -> list:
    base=key_norm(_strip_field_id(header))
    if base in option_index:
        return option_index[base]
    if not option_index:
        return []
    best=process.extractOne(base,list(option_index.keys()),scorer=fuzz.ratio)
    if best and best[1]>=92:
        return option_index[best[0]]
    return []


def _coerce_controlled(value, header: str, options: list):
    if not options:
        return None, "NO_OPTION_LIST_FOUND"
    n=key_norm(str(value))
    exact={key_norm(str(o)):o for o in options}
    if n in exact:
        return exact[n], "EXACT_OPTION"

    # Safe boolean derivations for yes/no controlled fields.
    opt_keys=set(exact)
    yes_key=next((k for k in opt_keys if k in {"si","yes","true"}),None)
    no_key=next((k for k in opt_keys if k in {"no","false"}),None)
    h=key_norm(header)
    if yes_key and no_key:
        neg_tokens={"no","false","none","sin","not supported","unsupported"}
        if n in neg_tokens or n.startswith("no ") or "sin " in n:
            return exact[no_key], "BOOLEAN_OPTION"
        # Presence of a concrete technology/certification can safely imply yes only for explicit boolean fields.
        h_compact=h.replace(" ","")
        if any(x in h for x in ["bluetooth","resistente al agua","water resistant","serial number","numero de serie"]) or any(x in h_compact for x in ["resistentealagua","waterresistant","serialnumber","numerodeserie"]):
            if str(value).strip():
                return exact[yes_key], "BOOLEAN_FROM_EVIDENCE"

    # Conservative normalized fuzzy match for textual controlled vocabularies.
    best=process.extractOne(n,list(exact.keys()),scorer=fuzz.ratio)
    if best and best[1]>=94:
        return exact[best[0]], "HIGH_CONFIDENCE_OPTION_MATCH"
    return None, "CONTROLLED_OPTION_NOT_FOUND"

def fill_excel(template: str, output: str, records: list[ProductRecord], overwrite: bool=False, trace_path: str|None=None):
    wb=load_workbook(template)
    option_index=_build_option_index(wb)
    trace=[]
    rejected=[]
    for ws in wb.worksheets:
        header_row,desc_row=_detect_header_and_description(ws)
        headers={c:ws.cell(header_row,c).value for c in range(1,ws.max_column+1) if ws.cell(header_row,c).value is not None}
        mapped={}
        contracts={}
        for c,h in headers.items():
            desc=ws.cell(desc_row,c).value if desc_row else None
            mapped[c]=map_header(str(h),str(desc) if desc not in (None,"") else None)
            key,_,field_class=mapped[c]
            contracts[c]=infer_contract(str(h),str(desc) if desc not in (None,"") else None,key,field_class)
        id_cols=[c for c,(k,conf,cls) in mapped.items() if k in IDENTITY_ALIASES and conf>=.75]
        if not id_cols: continue
        image_cols=[c for c,(k,conf,cls) in mapped.items() if k=="__image__" and conf>=.75]
        for row in range(header_row+1,ws.max_row+1):
            rec=None
            for c in id_cols:
                key,_,_=mapped[c]; val=ws.cell(row,c).value
                if val is None: continue
                for rr in records:
                    rv=getattr(rr.identity,key,None)
                    if rv and key_norm(str(rv))==key_norm(str(val)):
                        rec=rr; break
                if rec: break
            if not rec: continue
            for c,(key,mconf,field_class) in mapped.items():
                if key=="__image__" or mconf<.72: continue
                cell=ws.cell(row,c)
                if cell.value not in (None,"") and not overwrite: continue

                header_text = str(headers.get(c) or "")
                ext_id = _external_id(header_text)
                derived = derive_template_value(rec, header_text, ext_id)
                if derived.value not in (None,"") and derived.confidence >= .85:
                    ok,reason,gconf=validate_value(derived.value, contracts[c])
                    if ok:
                        cell.value=derived.value
                        trace.append({"sheet":ws.title,"cell":cell.coordinate,"attribute":"derived_marketplace_field","value":derived.value,"source":None,"field_class":"DERIVABLE","derivation_reason":derived.reason,"field_map_confidence":round(mconf,3),"value_confidence":derived.confidence,"guard":"OK"})
                    else:
                        rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header_text,"candidate":derived.value,"reason":reason,"stage":"derived"})
                    continue

                if not key: continue
                value,source,vconf,ev_attr,raw=_record_value(rec,key)
                if value in (None,""): continue
                ev=_candidate_evidence(rec,key)
                ev_attr=ev.attribute if ev else ev_attr
                raw=ev.raw_value if ev else raw
                if contracts[c].value_type == "controlled":
                    options=_find_options_for_field(option_index, header_text)
                    coerced, option_reason=_coerce_controlled(value, header_text, options)
                    if coerced is None:
                        rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header_text,"attribute":key,"candidate":value,"source":source,"reason":option_reason,"field_map_confidence":round(mconf,3),"value_confidence":vconf})
                        continue
                    value=coerced
                ok,reason,gconf=validate_value(value,contracts[c],evidence_attribute=ev_attr,evidence_raw=raw)
                if not ok:
                    rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header_text,"attribute":key,"candidate":value,"source":source,"reason":reason,"field_map_confidence":round(mconf,3),"value_confidence":vconf})
                    continue
                cell.value=value
                trace.append({"sheet":ws.title,"cell":cell.coordinate,"attribute":key,"value":value,"source":source,"field_class":field_class,"field_map_confidence":round(mconf,3),"value_confidence":vconf,"guard":"OK"})

            ranked=sorted(
                [i for i in rec.images if i.get("scope") in {"EXACT_VARIANT","EXACT_PRODUCT"} and i.get("confidence",0)>=.80],
                key=media_rank, reverse=True
            )
            seen=set()
            ranked=[i for i in ranked if not (i.get("url") in seen or seen.add(i.get("url")))]
            for idx,c in enumerate(image_cols):
                if idx>=len(ranked): break
                cell=ws.cell(row,c)
                if cell.value not in (None,"") and not overwrite: continue
                cell.value=ranked[idx]["url"]
                trace.append({"sheet":ws.title,"cell":cell.coordinate,"attribute":"product_image_url","value":ranked[idx]["url"],"source":ranked[idx]["url"],"media_scope":ranked[idx].get("scope"),"media_confidence":ranked[idx].get("confidence"),"guard":"EXACT_MEDIA_ONLY"})
    Path(output).parent.mkdir(parents=True,exist_ok=True)
    wb.save(output)
    report={"written":trace,"rejected":rejected,"summary":{"written_count":len(trace),"rejected_count":len(rejected)}}
    if trace_path:
        Path(trace_path).write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
    return report
