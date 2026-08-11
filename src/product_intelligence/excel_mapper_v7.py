from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

from .attribute_resolver import best_candidate
from .field_derivations import (
    derive_autonomy, derive_boolean, derive_connectivity, derive_controlled_color,
    derive_description, derive_headphone_type, derive_package_contents,
    derive_water_resistance, derive_features, derive_power_source, derive_segment,
)
from .marketplace_mapper import derive_template_value, media_rank
from .models import ProductRecord
from .normalize import ALIASES, canonical_key, key_norm
from .semantic_guard import infer_contract, is_placeholder, validate_value
from .template_intelligence import classify_field

IDENTITY_ALIASES={
    "mpn":["mpn","part number","part no","pn","codigo fabricante","código fabricante","manufacturer part number"],
    "ean":["ean","ean13","ean-13","codigo barras","código de barras","barcode"],
    "upc":["upc"],"gtin":["gtin"],"sku":["sku"],"model":["model","modelo"],"brand":["brand","marca"],
    "product_name":["nombre","product name","nombre del producto"],
}


def _strip_field_id(label:str)->str:
    return re.sub(r"#\s*[A-Za-z]*\d+", "", str(label)).strip()

def _external_id(label:str)->str|None:
    m=re.search(r"#\s*([A-Za-z]*\d+|\d+)",str(label)); return m.group(1) if m else None


def map_header(header:str,description:str|None=None):
    raw=_strip_field_id(header); h=key_norm(raw)
    cls,_,cconf=classify_field(header,description)
    if cls=="IMAGE": return "__image__",cconf,cls
    if cls=="SELLER_DATA": return None,cconf,cls
    if cls=="DERIVABLE": return None,cconf,cls
    pairs=[]
    for k,als in {**IDENTITY_ALIASES,**ALIASES}.items():
        pairs += [(key_norm(a),k) for a in [k,*als]]
    exact=next((k for a,k in pairs if a==h),None)
    if exact:return exact,1.0,cls
    best=process.extractOne(h,[a for a,_ in pairs],scorer=fuzz.ratio)
    if not best or best[1]<92:return None,(best[1]/100 if best else 0),cls
    return next(k for a,k in pairs if a==best[0]),best[1]/100,cls


def _detect_header_and_description(ws):
    best=(1,-1.0)
    for r in range(1,min(ws.max_row,25)+1):
        vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
        non=[str(v) for v in vals if v not in (None,'')]
        if not non: continue
        score=0.0
        for v in non:
            cls,_,_=classify_field(v)
            if canonical_key(_strip_field_id(v)): score+=3
            if re.search(r"#\s*[A-Za-z]*\d+",v):score+=2
            if cls in {"IMAGE","SELLER_DATA","DERIVABLE"}:score+=1
            if len(v)<80:score+=.2
        if score>best[1]:best=(r,score)
    header=best[0]; desc=None; ds=-1
    for r in range(max(1,header-3),header):
        vals=[str(ws.cell(r,c).value) for c in range(1,ws.max_column+1) if ws.cell(r,c).value not in (None,'')]
        avg=sum(map(len,vals))/len(vals) if vals else 0
        if avg>ds:desc,ds=r,avg
    return header,desc


def _build_option_index(wb)->dict[str,list]:
    idx={}
    for ws in wb.worksheets:
        # Generic one-column vocabulary sheets (e.g. Brands/Marcas, Categories/Categorías).
        if ws.max_column == 1 and ws.max_row >= 3:
            vals=[ws.cell(r,1).value for r in range(1,ws.max_row+1) if ws.cell(r,1).value not in (None,'')]
            if len(vals)>=3:
                idx[key_norm(ws.title)] = vals
                # If the first cell looks like a header, also index by it and exclude it from values.
                if vals and len(str(vals[0])) < 80:
                    idx.setdefault(key_norm(str(vals[0])), vals[1:] if len(vals)>3 else vals)
        for r in range(1,min(ws.max_row,5)+1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            compact=[v for v in vals if v not in (None,'') and len(str(v))<80]
            if len(compact)<2: continue
            populated=sum(1 for c,v in enumerate(vals,1) if v not in (None,'') and sum(1 for rr in range(r+1,min(ws.max_row,r+60)+1) if ws.cell(rr,c).value not in (None,''))>=2)
            if populated<2:continue
            for c,h in enumerate(vals,1):
                if h in (None,''):continue
                opts=[ws.cell(rr,c).value for rr in range(r+1,ws.max_row+1) if ws.cell(rr,c).value not in (None,'')]
                if opts:idx[key_norm(h)]=opts
            break
    return idx


def _find_options(idx,header):
    base=key_norm(_strip_field_id(header))
    if base in idx:return idx[base]
    alias_keys=[]
    if base in {'marca','brand'}: alias_keys=['marcas','brands','brand']
    elif 'categoria' in base or 'category' in base: alias_keys=['categorias','categories','category']
    for ak in alias_keys:
        if key_norm(ak) in idx:return idx[key_norm(ak)]
    best=process.extractOne(base,list(idx.keys()),scorer=fuzz.ratio) if idx else None
    return idx[best[0]] if best and best[1]>=94 else []


def _coerce_controlled(value,header,options):
    if not options:return None,"NO_OPTION_LIST_FOUND"
    exact={key_norm(str(o)):o for o in options}; n=key_norm(str(value))
    if n in exact:return exact[n],"EXACT_OPTION"
    # multi-select: every token must map exactly/near-exactly
    parts=[p.strip() for p in re.split(r"[,;|]",str(value)) if p.strip()]
    if len(parts)>1:
        out=[]
        for p in parts:
            np=key_norm(p)
            if np in exact: out.append(exact[np]);continue
            b=process.extractOne(np,list(exact.keys()),scorer=fuzz.ratio)
            if not b or b[1]<96:return None,"CONTROLLED_MULTI_OPTION_NOT_FOUND"
            out.append(exact[b[0]])
        return ", ".join(map(str,dict.fromkeys(out))),"MULTI_OPTION"
    b=process.extractOne(n,list(exact.keys()),scorer=fuzz.ratio)
    if b and b[1]>=96:return exact[b[0]],"HIGH_CONFIDENCE_OPTION_MATCH"
    return None,"CONTROLLED_OPTION_NOT_FOUND"


def _description_examples(desc:Any)->set[str]:
    text=str(desc or '')
    vals=set()
    for m in re.finditer(r"(?:Value|Valor)\s*:\s*([^\n]+)",text,re.I): vals.add(key_norm(m.group(1).strip()))
    for m in re.finditer(r"(?:Ej\.?|E\.g\.)\s*([^\n]+)",text,re.I): vals.add(key_norm(m.group(1).strip()))
    return {v for v in vals if v}


def _is_template_example(value,desc)->bool:
    if value in (None,''):return False
    n=key_norm(str(value))
    if is_placeholder(value):return True
    return n in _description_examples(desc)


def _match_record(records,row_values,mapped):
    best=None; best_score=0
    for rec in records:
        score=0; hard_conflict=False
        for c,(k,conf,cls) in mapped.items():
            if k not in IDENTITY_ALIASES:continue
            val=row_values.get(c)
            if val in (None,''):continue
            rv=getattr(rec.identity,k,None)
            if rv in (None,''):continue
            if key_norm(str(rv))==key_norm(str(val)):
                score += 4 if k in {'mpn','ean','upc','gtin'} else 1
            elif k in {'mpn','ean','upc','gtin'}:
                hard_conflict=True
        if hard_conflict:continue
        if score>best_score:best,best_score=rec,score
    return best if best_score>=1 else None


def _identity_value(rec,key):
    v=getattr(rec.identity,key,None)
    if v in (None,''):return None,0,None
    conf=.99 if rec.identity.match_level=='EXACT' else (.90 if rec.identity.match_level=='HIGH' else .0)
    return v,conf,"identity"


def _derived_for_field(rec,header,canonical,contract,options,ext_id):
    n=key_norm(_strip_field_id(header))
    # Existing V5 safe derivations.
    d=derive_template_value(rec,header,ext_id)
    if d.value not in (None,''):
        return d.value,d.confidence,d.reason,None,None
    if ext_id=='53' or n in {'descripcion','description'}:
        x=derive_description(rec);return x.value,x.confidence,x.reason,x.evidence_attribute,x.evidence_raw
    if ext_id in {'277497','277523'} or 'color' in n:
        x=derive_controlled_color(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='1568' or 'bluetooth' in n:
        x=derive_boolean(rec,'bluetooth');return x.value,x.confidence,x.reason,x.evidence_attribute,x.evidence_raw
    if ext_id=='36083' or 'resistentealagua' in n or 'waterresistance' in n:
        x=derive_water_resistance(rec,options);return x.value,x.confidence,x.reason,x.evidence_attribute,x.evidence_raw
    if ext_id=='1651' or 'conectividad' in n or 'connectivity' in n:
        x=derive_connectivity(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='1661' or 'tipodeauricular' in n:
        x=derive_headphone_type(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='1672' or 'autonomia' in n:
        x=derive_autonomy(rec);return x.value,x.confidence,x.reason,x.evidence_attribute,x.evidence_raw
    if ext_id=='1657' or 'caracteristicas' in n or 'features' in n:
        x=derive_features(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='1583' or 'alimentacion' in n or 'powersource' in n:
        x=derive_power_source(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='1536' or 'segmento' in n or 'segment' in n:
        x=derive_segment(rec,options);return x.value,x.confidence,x.reason,None,None
    if ext_id=='19' or 'contenidodelpaquete' in n:
        x=derive_package_contents(rec);return x.value,x.confidence,x.reason,x.evidence_attribute,x.evidence_raw
    return None,0,"no_derivation",None,None



def _media_key(url:str)->str:
    try:
        p=urlsplit(url)
        q=[(k,v) for k,v in parse_qsl(p.query,keep_blank_values=True) if key_norm(k) not in {"sw","sh","w","h","width","height","quality","q"}]
        return urlunsplit((p.scheme,p.netloc,p.path,urlencode(q),""))
    except Exception:
        return url

def fill_excel_v7(template:str,output:str,records:list[ProductRecord],overwrite:bool=False,trace_path:str|None=None):
    wb=load_workbook(template)
    options_idx=_build_option_index(wb)
    written=[];rejected=[];cleared_examples=[]
    for ws in wb.worksheets:
        hr,dr=_detect_header_and_description(ws)
        headers={c:ws.cell(hr,c).value for c in range(1,ws.max_column+1) if ws.cell(hr,c).value is not None}
        if not headers:continue
        mapped={};contracts={};descs={};opts={}
        for c,h in headers.items():
            desc=ws.cell(dr,c).value if dr else None;descs[c]=desc
            mapped[c]=map_header(str(h),str(desc) if desc not in (None,'') else None)
            k,_,cls=mapped[c]
            contracts[c]=infer_contract(str(h),str(desc) if desc else None,k,cls)
            opts[c]=_find_options(options_idx,str(h))
        id_cols=[c for c,(k,conf,cls) in mapped.items() if k in IDENTITY_ALIASES and conf>=.9]
        image_cols=[c for c,(k,conf,cls) in mapped.items() if k=='__image__']
        if not id_cols:continue
        for row in range(hr+1,ws.max_row+1):
            rowvals={c:ws.cell(row,c).value for c in headers}
            rec=_match_record(records,rowvals,mapped)
            if not rec or rec.identity.match_level == "CONFLICT":continue
            # Seller fields are never populated by scraping, but obvious examples shipped inside the
            # template are safe to remove so dummy prices/SKUs are not uploaded accidentally.
            for _c,(_k,_mc,_fc) in mapped.items():
                if _fc == "SELLER_DATA" and _is_template_example(ws.cell(row,_c).value, descs.get(_c)):
                    cleared_examples.append({"sheet":ws.title,"cell":ws.cell(row,_c).coordinate,"value":ws.cell(row,_c).value,"reason":"TEMPLATE_EXAMPLE_SELLER_DATA"})
                    ws.cell(row,_c).value=None
            if overwrite:
                # Overwrite mode means recompute product-derived cells from evidence. Seller/commercial
                # fields remain protected. This also removes stale outputs from older scraper versions.
                for _c,(_k,_mc,_fc) in mapped.items():
                    if _fc != "SELLER_DATA":
                        ws.cell(row,_c).value = None
            for c,(key,mconf,field_class) in mapped.items():
                if key=='__image__':continue
                cell=ws.cell(row,c);header=str(headers[c]);desc=descs[c]
                # Remove only values proven to be examples from the template instructions.
                if _is_template_example(cell.value,desc):
                    cleared_examples.append({"sheet":ws.title,"cell":cell.coordinate,"value":cell.value,"reason":"TEMPLATE_EXAMPLE"})
                    cell.value=None
                if field_class=='SELLER_DATA':
                    continue
                if cell.value not in (None,'') and not overwrite:continue
                ext=_external_id(header);options=opts[c]

                # Safe derivations first.
                value,vconf,reason,ev_attr,ev_raw=_derived_for_field(rec,header,key,contracts[c],options,ext)
                source=None
                if value not in (None,'') and vconf>=.85:
                    if contracts[c].value_type=='controlled':
                        cv,creason=_coerce_controlled(value,header,options)
                        if cv is None:
                            rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason})
                            continue
                        value=cv
                    ok,greason,gconf=validate_value(value,contracts[c],evidence_attribute=ev_attr,evidence_raw=ev_raw)
                    if ok:
                        cell.value=value
                        written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key or 'derived',"value":value,"source":source,"reason":reason,"confidence":round(vconf,3)})
                        continue
                    rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":greason})

                # Identity fields.
                if key in IDENTITY_ALIASES:
                    value,vconf,source=_identity_value(rec,key)
                    if value in (None,'') or vconf<.88:continue
                    if contracts[c].value_type=='controlled':
                        cv,creason=_coerce_controlled(value,header,options)
                        if cv is None:
                            rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason});continue
                        value=cv
                    ok,greason,_=validate_value(value,contracts[c])
                    if ok:
                        cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key,"value":value,"source":"identity","confidence":vconf});continue

                # Strict evidence resolver scans ALL clean evidence, including former additional_attributes.
                if key:
                    cand=best_candidate(rec,header,str(desc) if desc else None,key,contracts[c])
                    if cand:
                        value=cand.value
                        if contracts[c].value_type=='controlled':
                            cv,creason=_coerce_controlled(value,header,options)
                            if cv is None:
                                rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason,"attribute":cand.evidence.attribute});continue
                            value=cv
                        cell.value=value
                        written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key,"value":value,"source":cand.evidence.source_url,"evidence_attribute":cand.evidence.attribute,"confidence":round(cand.score,3),"reasons":cand.reasons})
            # Media only true product gallery assets and exact scope.
            ranked=sorted([i for i in rec.images if i.get('autofill_eligible',True) and i.get('role','product_gallery')=='product_gallery' and i.get('scope') in {'EXACT_VARIANT','EXACT_PRODUCT'} and i.get('confidence',0)>=.80],key=media_rank,reverse=True)
            seen=set(); dedup=[]
            for i in ranked:
                if not i.get('url'): continue
                k=_media_key(i['url'])
                if k in seen: continue
                seen.add(k); dedup.append(i)
            ranked=dedup
            for idx,c in enumerate(image_cols):
                if idx>=len(ranked):break
                cell=ws.cell(row,c)
                if cell.value not in (None,'') and not overwrite:continue
                cell.value=ranked[idx]['url']
                written.append({"sheet":ws.title,"cell":cell.coordinate,"header":str(headers[c]),"attribute":"product_image_url","value":ranked[idx]['url'],"source":ranked[idx]['url'],"scope":ranked[idx].get('scope'),"role":ranked[idx].get('role')})
    Path(output).parent.mkdir(parents=True,exist_ok=True);wb.save(output)
    report={"written":written,"rejected":rejected,"cleared_template_examples":cleared_examples,"summary":{"written_count":len(written),"rejected_count":len(rejected),"cleared_template_examples":len(cleared_examples)}}
    if trace_path:Path(trace_path).write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding='utf-8')
    return report
