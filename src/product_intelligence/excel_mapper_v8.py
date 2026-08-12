from __future__ import annotations
import json, re
from pathlib import Path
from typing import Any
from urllib.parse import urlparse,parse_qsl,urlencode,urlunparse
from openpyxl import load_workbook
from rapidfuzz import fuzz
from .models import ProductRecord
from .normalize import canonical_key,key_norm
from .attribute_resolver import best_candidate
from .semantic_guard import infer_contract,validate_value,is_placeholder
from .ai_enrichment import AIConfig,AIEnricher
from .template_intelligence import classify_field
from .template_contract import analyze_template_contract
from .field_derivations import derive_description,derive_connectivity,derive_headphone_type,derive_water_resistance,derive_power_source,derive_autonomy,derive_features,derive_segment,derive_boolean

IDENTITY_ALIASES={'brand','model','mpn','ean','upc','gtin','product name'}


def _strip_field_id(label:str)->str:return re.sub(r'#\s*[A-Za-z]*\d+','',str(label)).strip()
def _external_id(label:str)->str|None:
    m=re.search(r'#\s*([A-Za-z]*\d+)',str(label));return m.group(1) if m else None

def _detect_header_and_description(ws):
    best=None
    for r in range(1,min(ws.max_row,25)+1):
        vals=[str(ws.cell(r,c).value or '') for c in range(1,ws.max_column+1)]
        ids=sum(bool(re.search(r'#\s*[A-Za-z]*\d+',v)) for v in vals)
        mapped=sum(bool(canonical_key(_strip_field_id(v))) for v in vals if v)
        score=ids*4+mapped
        if best is None or score>best[1]:best=(r,score)
    hr=best[0]
    candidates=[]
    for r in range(max(1,hr-4),hr):
        vals=[str(ws.cell(r,c).value or '') for c in range(1,ws.max_column+1) if ws.cell(r,c).value not in (None,'')]
        avg=sum(map(len,vals))/len(vals) if vals else 0;candidates.append((avg,r))
    dr=max(candidates)[1] if candidates else None
    return hr,dr

def map_header(h,desc=None):
    field_class,ext_id,class_conf=classify_field(str(h),str(desc) if desc not in (None,'') else None)
    n=key_norm(_strip_field_id(h));ck=canonical_key(_strip_field_id(h))
    if field_class=='SELLER_DATA':return None,.99,'SELLER_DATA'
    if field_class=='IMAGE':return '__image__',.99,'IMAGE'
    if ck:return ck,1.0,field_class
    table=[
      (['nombre','name','titulo','title'],'product name'),(['marca','brand'],'brand'),(['modelo','model'],'model'),
      (['descripcion','description'],'description'),(['codigo de barras','barcode'],'ean'),(['conectividad','connectivity'],'connectivity'),
      (['bluetooth'],'bluetooth'),(['tipo de auricular','headphone type'],'headphone type'),(['resistente al agua','water resistance'],'water resistance'),
      (['alimentacion','power source'],'power source'),(['autonomia','battery life'],'battery life'),(['caracteristicas','features'],'features'),
      (['segmento','segment'],'segment'),(['tipo de salida','output type'],'output type'),(['contenido del paquete','package contents'],'package contents'),
      (['ancho del paquete','package width'],'package width'),(['largo del paquete','package length'],'package length'),(['alto del paquete','package height'],'package height'),(['peso del paquete','package weight'],'package weight'),
      (['alto','height'],'height'),(['ancho','width'],'width'),(['largo','length'],'length'),(['dimensiones','dimensions'],'dimensions'),(['potencia','power'],'power'),
      (['garantia del producto','product warranty'],'product warranty'),(['pais de produccion','country of production'],'country of origin'),(['color'],'color'),
    ]
    for toks,k in table:
        if any(key_norm(x) in n for x in toks):return k,.92,field_class
    return None,class_conf,field_class

def _is_template_example(value,desc):
    if value in (None,''):return False
    v=key_norm(str(value));d=key_norm(str(desc or ''))
    if is_placeholder(value):return True
    if v in {'esto es un parrafo','1234567890','abc 1000 202','999 999 99'}:return True
    nums=re.findall(r'(?<!\d)(\d+(?:[.,]\d+)?)(?!\d)',str(desc or ''))
    if str(value).strip() in nums and ('value' in d or 'example' in d or 'ejemplo' in d):return True
    if 'e.g.' in d or 'ejemplo' in d:
        if v and v in d:return True
    return False

def _identity_value(rec,key):
    attr={'product name':'product_name'}.get(key,key)
    val=getattr(rec.identity,attr,None)
    conf=.99 if rec.identity.match_level=='EXACT' else (.90 if rec.identity.match_level=='HIGH' else float(rec.identity.confidence or 0))
    return val,conf,'identity'

def _option_keys(label):
    s=str(label or '');return {key_norm(s),key_norm(_strip_field_id(s))}-{''}
def _build_option_index(wb):
    idx={}
    for ws in wb.worksheets:
        for r in range(1,min(ws.max_row,8)+1):
            heads=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            populated=0
            for c,h in enumerate(heads,1):
                if h in (None,''):continue
                below=[ws.cell(rr,c).value for rr in range(r+1,min(ws.max_row,r+80)+1)]
                if sum(x not in (None,'') for x in below)>=2:populated+=1
            if populated<2:continue
            for c,h in enumerate(heads,1):
                if h in (None,''):continue
                vals=[ws.cell(rr,c).value for rr in range(r+1,ws.max_row+1) if ws.cell(rr,c).value not in (None,'')]
                if vals:
                    for k in _option_keys(h):idx[k]=vals
            break
        if ws.max_column==1 and ws.max_row>=2:
            vals=[ws.cell(r,1).value for r in range(1,ws.max_row+1) if ws.cell(r,1).value not in (None,'')]
            if len(vals)>=2:
                idx[key_norm(ws.title)]=vals
                if len(str(vals[0]))<100:
                    for k in _option_keys(vals[0]):idx[k]=vals[1:]
    return idx
def _find_options(idx,label):
    for k in _option_keys(label):
        if k in idx:return idx[k]
    base=key_norm(_strip_field_id(label));aliases=[]
    if base in {'marca','brand'}:aliases=['marcas','brands','brand']
    elif 'categoria' in base or 'category' in base:aliases=['categorias','categories','category']
    for a in aliases:
        if key_norm(a) in idx:return idx[key_norm(a)]
    return []
def _coerce_controlled(value,header,options):
    if not options:return None,'NO_ALLOWED_OPTIONS'
    vals=value if isinstance(value,list) else [x.strip() for x in re.split(r'[,;|]',str(value)) if x.strip()]
    out=[]
    for v in vals:
        nv=key_norm(str(v));exact=next((o for o in options if key_norm(str(o))==nv),None)
        if exact is None:
            exact=next((o for o in options if key_norm(str(o)).rstrip('s')==nv.rstrip('s') and len(nv)>=4),None)
        if exact is None:return None,f'VALUE_NOT_IN_ALLOWED_OPTIONS:{v}'
        if exact not in out:out.append(exact)
    return ', '.join(map(str,out)),'OK'
def _derived_for_field(rec,header,description,canonical,contract,options,ext_id):
    h=key_norm(header);intent=key_norm(f'{header} {description or ""} {getattr(contract,"semantic","") or ""}')
    d=None
    if ext_id=='53' or canonical=='description' or h in {'descripcion','description'}:d=derive_description(rec)
    elif 'bluetooth' in intent:d=derive_boolean(rec,'bluetooth')
    elif 'resistente al agua' in intent or 'water resistance' in intent:d=derive_water_resistance(rec,options)
    elif canonical=='connectivity' or 'conectividad' in intent or 'connectivity' in intent:d=derive_connectivity(rec,options)
    elif canonical=='headphone type' or 'tipo de auricular' in intent or 'headphone type' in intent:d=derive_headphone_type(rec,options)
    elif canonical=='power source' or 'alimentacion' in intent or 'power source' in intent:d=derive_power_source(rec,options)
    elif canonical=='battery life' or 'autonomia' in intent or 'battery life' in intent:d=derive_autonomy(rec)
    elif canonical=='features' or 'caracteristicas' in intent or 'features' in intent:d=derive_features(rec,options)
    elif canonical=='segment' or 'segmento' in intent:d=derive_segment(rec,options)
    if not d:return None,0,'',None,None
    ev_attr=None;ev_raw=None
    if d.evidence_attribute:ev_attr=d.evidence_attribute
    if d.evidence_raw is not None:ev_raw=d.evidence_raw
    return d.value,d.confidence,d.reason,ev_attr,ev_raw
def _match_record(records,rowvals,mapped):
    best=None
    for rec in records:
        score=0
        for c,(k,conf,cls) in mapped.items():
            v=rowvals.get(c)
            if v in (None,''):continue
            nv=key_norm(str(v))
            for attr in ['mpn','ean','upc','gtin','model','product_name']:
                rv=getattr(rec.identity,attr,None)
                if rv and (nv==key_norm(str(rv)) or (len(nv)>5 and nv in key_norm(str(rv)))):
                    score+=100 if attr in {'mpn','ean','upc','gtin'} else 20
        if best is None or score>best[0]:best=(score,rec)
    return best[1] if best and best[0]>0 else None
def _media_key(url):
    p=urlparse(str(url));q=[(k,v) for k,v in parse_qsl(p.query,keep_blank_values=True) if k.lower() not in {'sw','sh','w','h','width','height','quality','q','format'}]
    return urlunparse((p.scheme.lower(),p.netloc.lower(),p.path,'',urlencode(q),''))
def media_rank(item):
    s=0
    if item.get('scope')=='EXACT_VARIANT':s+=100
    elif item.get('scope')=='EXACT_PRODUCT':s+=80
    if item.get('role')=='product_gallery':s+=25
    src=key_norm(item.get('source') or '')
    if 'jsonld product image' in src:s+=15
    if 'zoom' in src or 'large' in src:s+=8
    if 'og image' in src:s+=3
    source_class_rank={'manufacturer':3,'secondary':2,'marketplace':1}.get(str(item.get('source_class') or ''),0)
    return (s,source_class_rank,float(item.get('confidence',0)))

def fill_excel_v8(template,output,records,overwrite=False,trace_path=None,ai_config:AIConfig|None=None,row_assignments:dict[tuple[str,int],ProductRecord]|None=None):
    ai=AIEnricher(ai_config or AIConfig())
    row_assignments=row_assignments or {}
    wb=load_workbook(template)
    template_plan=analyze_template_contract(template)
    target_lookup={(f["sheet"],f["column"]):f for sh in template_plan.get("sheets",[]) for f in sh.get("fields",[])}
    options_idx=_build_option_index(wb)
    written=[];rejected=[];cleared_examples=[]
    for ws in wb.worksheets:
        hr,dr=_detect_header_and_description(ws)
        headers={c:ws.cell(hr,c).value for c in range(1,ws.max_column+1) if ws.cell(hr,c).value is not None}
        if not headers:continue
        mapped={};contracts={};descs={};opts={}
        for c,h in headers.items():
            desc=ws.cell(dr,c).value if dr else None;descs[c]=desc
            mapped[c]=map_header(str(h),str(desc) if desc not in (None,'') else None)
            k,_,cls=mapped[c]
            contracts[c]=infer_contract(str(h),str(desc) if desc else None,k,cls)
            opts[c]=_find_options(options_idx,str(h))
        id_cols=[c for c,(k,conf,cls) in mapped.items() if k in IDENTITY_ALIASES and conf>=.9]
        image_cols=[c for c,(k,conf,cls) in mapped.items() if k=='__image__']
        if not id_cols:continue
        assigned_rows=[r for (sheet_name,r) in row_assignments if sheet_name==ws.title]
        last_row=max([ws.max_row, *assigned_rows]) if assigned_rows else ws.max_row
        for row in range(hr+1,last_row+1):
            rowvals={c:ws.cell(row,c).value for c in headers}
            rec=row_assignments.get((ws.title,row)) or _match_record(records,rowvals,mapped)
            if not rec or rec.identity.match_level == "CONFLICT":continue
            for _c,(_k,_mc,_fc) in mapped.items():
                if _fc == "SELLER_DATA" and _is_template_example(ws.cell(row,_c).value, descs.get(_c)):
                    cleared_examples.append({"sheet":ws.title,"cell":ws.cell(row,_c).coordinate,"value":ws.cell(row,_c).value,"reason":"TEMPLATE_EXAMPLE_SELLER_DATA"})
                    ws.cell(row,_c).value=None
            if overwrite:
                for _c,(_k,_mc,_fc) in mapped.items():
                    _plan=target_lookup.get((ws.title,_c),{})
                    if _fc != "SELLER_DATA" or _plan.get("role")=="SCRAPE_TARGET":
                        ws.cell(row,_c).value = None
            for c,(key,mconf,field_class) in mapped.items():
                if key=='__image__':continue
                cell=ws.cell(row,c);header=str(headers[c]);desc=descs[c]
                if _is_template_example(cell.value,desc):
                    cleared_examples.append({"sheet":ws.title,"cell":cell.coordinate,"value":cell.value,"reason":"TEMPLATE_EXAMPLE"})
                    cell.value=None
                plan_field=target_lookup.get((ws.title,c),{})
                if (field_class=='SELLER_DATA' and plan_field.get("role")!="SCRAPE_TARGET") or plan_field.get("role") in {"SELLER_INPUT","MARKETPLACE_INPUT","DERIVED_OUTPUT"}:
                    continue
                if cell.value not in (None,'') and not overwrite:continue
                ext=_external_id(header);options=opts[c]
                if ai.config.enabled and (ext=='53' or key_norm(_strip_field_id(header)) in {'descripcion','description'}) and rec.identity.match_level not in {'CONFLICT','LOW'}:
                    suggestion=ai.suggest(rec,header,str(desc) if desc else None,None,language='es')
                    if suggestion and float(suggestion.get('confidence') or 0)>=.84:
                        value=suggestion['value'];evs=suggestion.get('evidence') or []
                        ev_attr='; '.join(str(x.get('attribute','')) for x in evs[:6]);ev_raw='; '.join(str(x.get('value','')) for x in evs[:6])
                        ok,greason,_=validate_value(value,contracts[c],evidence_attribute=ev_attr,evidence_raw=ev_raw)
                        if ok:
                            cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":"ai_description","value":value,"source":"AI over validated evidence","reason":suggestion.get("reason"),"confidence":round(float(suggestion.get('confidence') or 0),3),"evidence_ids":suggestion.get("evidence_ids")});continue
                        rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":greason,"layer":"ai_description"})
                value,vconf,reason,ev_attr,ev_raw=_derived_for_field(rec,header,str(desc) if desc else None,key,contracts[c],options,ext)
                source=None
                if value not in (None,'') and vconf>=.85:
                    if contracts[c].value_type=='controlled':
                        cv,creason=_coerce_controlled(value,header,options)
                        if cv is None:
                            rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason});continue
                        value=cv
                    ok,greason,gconf=validate_value(value,contracts[c],evidence_attribute=ev_attr,evidence_raw=ev_raw)
                    if ok:
                        cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key or 'derived',"value":value,"source":source,"reason":reason,"confidence":round(vconf,3)});continue
                    rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":greason})
                if key in IDENTITY_ALIASES:
                    value,vconf,source=_identity_value(rec,key)
                    if value in (None,'') or vconf<.88:continue
                    if contracts[c].value_type=='controlled':
                        cv,creason=_coerce_controlled(value,header,options)
                        if cv is None:
                            rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason});continue
                        value=cv
                    ok,greason,_=validate_value(value,contracts[c])
                    if ok:
                        cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key,"value":value,"source":"identity","confidence":vconf});continue
                if key:
                    cand=best_candidate(rec,header,str(desc) if desc else None,key,contracts[c])
                    if cand:
                        value=cand.value
                        if contracts[c].value_type=='controlled':
                            cv,creason=_coerce_controlled(value,header,options)
                            if cv is None:
                                rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":creason,"attribute":cand.evidence.attribute});continue
                            value=cv
                        cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key,"value":value,"source":cand.evidence.source_url,"evidence_attribute":cand.evidence.attribute,"confidence":round(cand.score,3),"reasons":cand.reasons});continue
                if ai.config.enabled and rec.identity.match_level not in {'CONFLICT','LOW'}:
                    suggestion=ai.suggest(rec,header,str(desc) if desc else None,options if contracts[c].value_type=='controlled' else None,language='es')
                    if suggestion and float(suggestion.get('confidence') or 0)>=.82:
                        value=suggestion['value'];evs=suggestion.get('evidence') or []
                        ev_attr='; '.join(str(x.get('attribute','')) for x in evs[:4]);ev_raw='; '.join(str(x.get('value','')) for x in evs[:4])
                        ok,greason,_=validate_value(value,contracts[c],evidence_attribute=ev_attr,evidence_raw=ev_raw)
                        if ok:
                            cell.value=value;written.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"attribute":key or 'ai_derived',"value":value,"source":"AI over validated evidence","reason":suggestion.get('reason'),"confidence":round(float(suggestion.get('confidence') or 0),3),"evidence_ids":suggestion.get('evidence_ids')});continue
                        rejected.append({"sheet":ws.title,"cell":cell.coordinate,"header":header,"candidate":value,"reason":greason,"layer":"ai"})
            ranked=sorted([i for i in rec.images if i.get('autofill_eligible',True) and i.get('role','product_gallery')=='product_gallery' and i.get('scope') in {'EXACT_VARIANT','EXACT_PRODUCT'} and i.get('confidence',0)>=.80],key=media_rank,reverse=True)
            seen=set();dedup=[]
            for i in ranked:
                u=i.get('url')
                if not u:continue
                k=_media_key(u)
                if k in seen:continue
                seen.add(k);dedup.append(i)
            for idx,c in enumerate(image_cols):
                if idx>=len(dedup):break
                cell=ws.cell(row,c)
                if cell.value not in (None,'') and not overwrite:continue
                cell.value=dedup[idx]['url'];written.append({"sheet":ws.title,"cell":cell.coordinate,"header":str(headers[c]),"attribute":"product_image_url","value":dedup[idx]['url'],"source":dedup[idx].get('source_page') or dedup[idx]['url'],"source_class":dedup[idx].get('source_class'),"scope":dedup[idx].get('scope'),"role":dedup[idx].get('role')})
    Path(output).parent.mkdir(parents=True,exist_ok=True);wb.save(output)
    report={"written":written,"rejected":rejected,"cleared_template_examples":cleared_examples,"summary":{"written_count":len(written),"rejected_count":len(rejected),"cleared_template_examples":len(cleared_examples)}}
    if trace_path:Path(trace_path).write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding='utf-8')
    return report
