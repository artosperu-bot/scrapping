from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .excel_intake import WorkbookIntakeResult, analyze_workbook_intake as _base_analyze_workbook_intake
from .models import ProductIdentity
from .normalize import key_norm


_PART_NUMBER_RE = re.compile(
    r"(?im)(?:^|[\n\r;])\s*(?:PART\s*NUMBER|PARTNUMBER|MPN|N[UÚ]MERO\s+DE\s+PARTE)\s*[:#=-]\s*([A-Z0-9][A-Z0-9._/-]{2,79})\b"
)
_GTIN_RE = re.compile(
    r"(?im)(?:^|[\n\r;])\s*(?:EAN\s*/\s*UPC|EAN|UPC|GTIN)\s*[:#=-]\s*([0-9][0-9\s-]{6,20}[0-9])\b"
)


def _clean(value) -> str:
    return str(value or "").strip()


def _compact(value) -> str:
    return re.sub(r"[^a-z0-9]", "", key_norm(_clean(value)))


def _looks_like_field_dictionary(ws) -> bool:
    """Detect vertical schema/reference sheets by structure, not marketplace/sheet name."""
    matches = 0
    comparable = 0
    end = min(ws.max_row, 40)
    for row in range(2, end + 1):
        left = _clean(ws.cell(row, 1).value)
        right = _clean(ws.cell(row, 2).value)
        if not left or not right or len(left) > 100 or len(right) > 100:
            continue
        comparable += 1
        a = _compact(left)
        b = _compact(right)
        if a and b and (a == b or (len(a) >= 5 and a in b) or (len(b) >= 5 and b in a)):
            matches += 1
    return comparable >= 4 and matches >= 4 and (matches / comparable) >= 0.35


def _explicit_identity_from_row(ws, row: int) -> tuple[str | None, str | None]:
    part_number = None
    gtin = None
    for col in range(1, ws.max_column + 1):
        text = _clean(ws.cell(row, col).value)
        if not text:
            continue
        if part_number is None:
            match = _PART_NUMBER_RE.search(text)
            if match:
                part_number = match.group(1).strip()
        if gtin is None:
            match = _GTIN_RE.search(text)
            if match:
                digits = re.sub(r"\D", "", match.group(1))
                if len(digits) in {8, 12, 13, 14}:
                    gtin = digits
        if part_number and gtin:
            break
    return part_number, gtin


def _enrich_product(product, ws) -> None:
    part_number, gtin = _explicit_identity_from_row(ws, product.row)
    if not part_number and not gtin:
        return

    values = product.identity.model_dump()
    if part_number and not values.get("mpn"):
        values["mpn"] = part_number
    if gtin and not (values.get("gtin") or values.get("ean") or values.get("upc")):
        values["gtin"] = gtin
    product.identity = ProductIdentity(**values)

    if part_number:
        product.audit["part_number_content"] = part_number
        product.audit["identity_selected"] = part_number
        product.audit["identity_type"] = "PART_NUMBER_FROM_CONTENT"
    if gtin:
        product.audit["gtin_content"] = gtin


def analyze_workbook_intake(template: str) -> WorkbookIntakeResult:
    """Return product rows after structural sheet filtering and explicit identity enrichment."""
    result = _base_analyze_workbook_intake(template)
    path = Path(template)
    wb = load_workbook(path, data_only=False, read_only=False)
    schema_sheets = {ws.title for ws in wb.worksheets if _looks_like_field_dictionary(ws)}

    if schema_sheets:
        result.products = [product for product in result.products if product.sheet not in schema_sheets]
        for audit in result.sheets:
            if audit.sheet in schema_sheets:
                audit.accepted = False
                audit.rejection_reason = "STRUCTURAL_SCHEMA_SHEET"
                audit.rows.clear()

    worksheets = {ws.title: ws for ws in wb.worksheets}
    for product in result.products:
        ws = worksheets.get(product.sheet)
        if ws is not None:
            _enrich_product(product, ws)
    return result
