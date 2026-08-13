from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

SELLER_SKU_LABELS = {"seller sku", "sku vendedor", "sku del vendedor"}


def _norm(value) -> str:
    text = str(value or "").lower()
    text = re.sub(r"#\s*[A-Za-z]*\d+", "", text)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _apply(output, records, row_assignments=None) -> list[dict]:
    from .excel_mapper_v8 import _detect_header_and_description, _match_record, map_header

    row_assignments = row_assignments or {}
    wb = load_workbook(output)
    written = []
    normalized_labels = {_norm(x) for x in SELLER_SKU_LABELS}
    for ws in wb.worksheets:
        header_row, _description_row = _detect_header_and_description(ws)
        headers = {col: ws.cell(header_row, col).value for col in range(1, ws.max_column + 1) if ws.cell(header_row, col).value is not None}
        sku_cols = [col for col, header in headers.items() if _norm(header) in normalized_labels]
        if not sku_cols:
            continue
        mapped = {col: map_header(str(header)) for col, header in headers.items()}
        assigned_rows = [row for (sheet_name, row) in row_assignments if sheet_name == ws.title]
        last_row = max([ws.max_row, *assigned_rows]) if assigned_rows else ws.max_row
        for row in range(header_row + 1, last_row + 1):
            row_values = {col: ws.cell(row, col).value for col in headers}
            rec = row_assignments.get((ws.title, row)) or _match_record(records, row_values, mapped)
            if not rec or rec.identity.match_level == "CONFLICT" or not rec.identity.mpn:
                continue
            for col in sku_cols:
                cell = ws.cell(row, col)
                cell.value = str(rec.identity.mpn)
                written.append({
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "header": str(headers[col]),
                    "field": "seller_sku",
                    "value": str(rec.identity.mpn),
                    "status": "FOUND_DERIVED",
                    "confidence": 1.0,
                    "source": "identity.mpn",
                    "reason": "seller_sku_defaults_to_part_number",
                })
    if written:
        Path(output).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
    return written


def install() -> None:
    """Keep seller fields protected while making seller SKU deterministic at write time."""
    from . import excel_mapper_v8

    if getattr(excel_mapper_v8.fill_excel_v8, "_seller_sku_default_wrapped", False):
        return
    original = excel_mapper_v8.fill_excel_v8

    def wrapped(template, output, records, overwrite=False, trace_path=None, row_assignments=None):
        report = original(template, output, records, overwrite=overwrite, trace_path=trace_path, row_assignments=row_assignments)
        seller_written = _apply(output, records, row_assignments=row_assignments)
        if seller_written:
            report.setdefault("written", []).extend(seller_written)
            summary = report.setdefault("summary", {})
            summary["written_count"] = len(report["written"])
        return report

    wrapped._seller_sku_default_wrapped = True
    excel_mapper_v8.fill_excel_v8 = wrapped
