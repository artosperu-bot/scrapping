from __future__ import annotations

import argparse
import json
from contextlib import ExitStack
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from . import batch
from .excel_mapper_v8 import fill_excel_v8
from .models import Evidence, ProductIdentity, ProductRecord
from .source_strategy import SourceStrategy


def _identity() -> ProductIdentity:
    return ProductIdentity(
        brand="Example",
        model="Example Model Wireless",
        mpn="EX-100-WL",
        confidence=.99,
        match_level="EXACT",
        identifiers_confirmed=["mpn"],
    )


def _record(
    fields,
    *,
    source_url: str,
    source_type: str,
    relationship: str = "EXACT_MODEL",
    scope: str = "MODEL",
    policy_allowed: bool = True,
) -> ProductRecord:
    identity = _identity()
    evidence = [
        Evidence(
            attribute=field,
            raw_value=value,
            normalized_value=value,
            source_url=source_url,
            source_type=source_type,
            extraction_method="pdf_native" if "pdf" in source_type else "jsonld",
            match_level="EXACT",
            confidence=.98,
            identity_status="EXACT",
            authority="manufacturer",
            policy_allowed=policy_allowed,
            document_relationship=relationship,
            document_scope=scope,
        )
        for field, value in fields
    ]
    return ProductRecord(
        identity=identity,
        evidence=evidence,
        sources=[source_url],
        fetch={
            "source_class": "manufacturer",
            "final_url": source_url,
            "source_decision": {
                "page_type": "DOCUMENT" if "pdf" in source_type else "PRODUCT",
                "identity": "EXACT",
                "authority": "manufacturer",
                "material_allowed": True,
            },
        },
    )


def _make_color_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws["A1"] = "MPN"
    ws["B1"] = "Color"
    ws["A2"] = "EX-100-WL"
    wb.save(path)


def _write_record_to_color_excel(template: Path, output: Path, record: ProductRecord) -> None:
    fill_excel_v8(
        str(template),
        str(output),
        [],
        overwrite=False,
        row_assignments={("Products", 2): record},
    )


def _exact_sku_web_to_excel(root: Path) -> dict:
    scenario_dir = root / "exact-sku-web"
    scenario_dir.mkdir(parents=True, exist_ok=True)
    source_url = "https://manufacturer.test/example-product"
    template = scenario_dir / "template.xlsx"
    output = scenario_dir / "verified.xlsx"
    _make_color_template(template)

    exact_record = _record(
        (("color", "Black"),),
        source_url=source_url,
        source_type="manufacturer_web",
        relationship="EXACT_SKU",
        scope="SKU",
    )

    class FakePipeline:
        def process_url(self, *args, **kwargs):
            return exact_record

    item = batch.BatchItem(
        row=2,
        sheet="Products",
        identity=_identity(),
        source_url=source_url,
    )
    template_plan = {
        "scrape_semantics": ["color"],
        "media_slots": 0,
        "summary": {"scrape_targets": 1},
    }

    with ExitStack() as stack:
        stack.enter_context(patch.object(batch, "ProductPipeline", FakePipeline))
        stack.enter_context(patch.object(batch, "search_web", lambda *a, **k: []))
        stack.enter_context(patch.object(batch, "search_web_for_fields", lambda *a, **k: []))
        stack.enter_context(patch.object(batch, "discover_product_documents", lambda *a, **k: []))
        record = batch.scrape_item(
            item,
            str(scenario_dir / "json"),
            template_plan=template_plan,
            source_strategy=SourceStrategy(web=True, pdf=True, ocr=False, mistral=False),
        )

    if record is None:
        return {"scenario": "EXACT_SKU_WEB_TO_EXCEL", "status": "FAIL", "reason": "NO_RECORD"}

    _write_record_to_color_excel(template, output, record)
    value = load_workbook(output, data_only=False)["Products"]["B2"].value
    audit = (record.evidence_graph or {}).get("smart_orchestrator") or {}
    status = "PASS" if value == "Black" and audit.get("resolved_fields") == ["color"] else "FAIL"
    return {
        "scenario": "EXACT_SKU_WEB_TO_EXCEL",
        "status": status,
        "category": audit.get("category"),
        "resolved_fields": list(audit.get("resolved_fields") or []),
        "missing_fields": list(audit.get("missing_fields") or []),
        "written_value": value,
        "output_excel": str(output.resolve()),
        "output_exists": output.is_file(),
    }


def _pdf_zero_web_fallback(root: Path) -> dict:
    scenario_dir = root / "pdf-zero-fallback"
    scenario_dir.mkdir(parents=True, exist_ok=True)
    web_url = "https://manufacturer.test/example-driver"
    candidate = SimpleNamespace(url=web_url, likely_official=True, score=1.0)
    broad_web_calls = 0
    targeted_web_calls = 0

    def broad_search(*args, **kwargs):
        nonlocal broad_web_calls
        broad_web_calls += 1
        return []

    def targeted_search(_identity, fields, **kwargs):
        nonlocal targeted_web_calls
        targeted_web_calls += 1
        return [candidate] if "driver_size" in fields else []

    class FakePipeline:
        def process_url(self, *args, **kwargs):
            return _record(
                (("driver_size", "40 mm"),),
                source_url=web_url,
                source_type="manufacturer_web",
                relationship="EXACT_MODEL",
                scope="MODEL",
            )

    item = batch.BatchItem(row=2, sheet="Products", identity=_identity())
    template_plan = {
        "scrape_semantics": ["driver_size"],
        "media_slots": 0,
        "summary": {"scrape_targets": 1},
    }

    with ExitStack() as stack:
        stack.enter_context(patch.object(batch, "ProductPipeline", FakePipeline))
        stack.enter_context(patch.object(batch, "search_web", broad_search))
        stack.enter_context(patch.object(batch, "search_web_for_fields", targeted_search))
        stack.enter_context(patch.object(batch, "discover_product_documents", lambda *a, **k: []))
        record = batch.scrape_item(
            item,
            str(scenario_dir / "json"),
            template_plan=template_plan,
            source_strategy=SourceStrategy(web=True, pdf=True, ocr=False, mistral=False),
        )

    audit = (record.evidence_graph or {}).get("smart_orchestrator") if record else {}
    resolved = list((audit or {}).get("resolved_fields") or [])
    status = "PASS" if record and broad_web_calls == 0 and targeted_web_calls == 1 and resolved == ["driver_size"] else "FAIL"
    return {
        "scenario": "PDF_ZERO_WEB_FALLBACK",
        "status": status,
        "pdf_documents_found": 0,
        "broad_web_calls": broad_web_calls,
        "targeted_web_calls": targeted_web_calls,
        "resolved_fields": resolved,
        "missing_fields": list((audit or {}).get("missing_fields") or []),
        "stop_reason": (audit or {}).get("stop_reason"),
    }


def _sibling_write_barrier(root: Path) -> dict:
    scenario_dir = root / "sibling-write-barrier"
    scenario_dir.mkdir(parents=True, exist_ok=True)
    template = scenario_dir / "template.xlsx"
    output = scenario_dir / "blocked.xlsx"
    _make_color_template(template)

    sibling = _record(
        (("color", "Red"),),
        source_url="https://manufacturer.test/sibling-product",
        source_type="manufacturer_web",
        relationship="SIBLING_VARIANT",
        scope="SKU",
        policy_allowed=True,
    )
    _write_record_to_color_excel(template, output, sibling)
    value = load_workbook(output, data_only=False)["Products"]["B2"].value
    status = "PASS" if value is None else "FAIL"
    return {
        "scenario": "SIBLING_WRITE_BARRIER",
        "status": status,
        "written_value": value,
        "output_excel": str(output.resolve()),
        "output_exists": output.is_file(),
    }


def run_smoke(output_dir: str | Path) -> dict:
    root = Path(output_dir)
    root.mkdir(parents=True, exist_ok=True)
    scenarios = [
        _exact_sku_web_to_excel(root),
        _pdf_zero_web_fallback(root),
        _sibling_write_barrier(root),
    ]
    status = "PASS" if all(row.get("status") == "PASS" for row in scenarios) else "FAIL"
    report = {
        "entrypoint": "SMART_PACKAGED_E2E",
        "status": status,
        "scenarios": scenarios,
    }
    report_path = root / "smart-packaged-smoke.json"
    report["report_path"] = str(report_path.resolve())
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main(argv=None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--smart-e2e-smoke", action="store_true")
    parser.add_argument("--output-dir", default="smart-packaged-evidence")
    args = parser.parse_args(argv)
    report = run_smoke(args.output_dir)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["status"] == "PASS" else 2


if __name__ == "__main__":
    raise SystemExit(main())
