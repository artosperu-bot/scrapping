from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

from .normalize import key_norm
from .template_intelligence import analyze_matrix


ROLE_IDENTITY = "IDENTITY"
ROLE_SCRAPE = "SCRAPE_TARGET"
ROLE_DERIVED = "DERIVED_OUTPUT"
ROLE_MEDIA = "MEDIA_TARGET"
ROLE_SELLER = "SELLER_INPUT"
ROLE_MARKETPLACE = "MARKETPLACE_INPUT"
ROLE_REVIEW = "REVIEW_REQUIRED"


@dataclass
class TemplateTarget:
    sheet: str
    column: int
    label: str
    description: str | None
    group: str | None
    external_id: str | None
    canonical: str | None
    semantic: str | None
    field_class: str
    role: str
    required: bool
    value_type: str
    options: list[Any]
    scrape: bool
    reason: str

    def to_dict(self) -> dict:
        return asdict(self)


def _strip_field_id(label: str) -> str:
    return re.sub(r"#\s*[A-Za-z]*\d+", "", str(label)).strip()


def _build_option_index(wb) -> dict[str, list[Any]]:
    idx: dict[str, list[Any]] = {}
    for ws in wb.worksheets:
        if ws.max_column == 1 and ws.max_row >= 2:
            vals = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value not in (None, "")]
            if len(vals) >= 2:
                idx[key_norm(ws.title)] = vals
                if len(str(vals[0])) < 80:
                    idx.setdefault(key_norm(str(vals[0])), vals[1:])
        for r in range(1, min(ws.max_row, 6) + 1):
            headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            usable = [v for v in headers if v not in (None, "") and len(str(v)) < 120]
            if len(usable) < 2:
                continue
            populated = 0
            for c, h in enumerate(headers, 1):
                if h in (None, ""):
                    continue
                below = [ws.cell(rr, c).value for rr in range(r + 1, min(ws.max_row, r + 80) + 1)]
                if sum(1 for x in below if x not in (None, "")) >= 2:
                    populated += 1
            if populated < 2:
                continue
            for c, h in enumerate(headers, 1):
                if h in (None, ""):
                    continue
                opts = [ws.cell(rr, c).value for rr in range(r + 1, ws.max_row + 1) if ws.cell(rr, c).value not in (None, "")]
                if opts:
                    idx[key_norm(str(h))] = opts
            break
    return idx


def _find_options(idx: dict[str, list[Any]], label: str) -> list[Any]:
    base = key_norm(_strip_field_id(label))
    if base in idx:
        return idx[base]
    aliases: list[str] = []
    if base in {"marca", "brand"}:
        aliases = ["marcas", "brands", "brand"]
    elif "categoria" in base or "category" in base:
        aliases = ["categorias", "categories", "category"]
    for alias in aliases:
        if key_norm(alias) in idx:
            return idx[key_norm(alias)]
    best = process.extractOne(base, list(idx), scorer=fuzz.ratio) if idx else None
    return idx[best[0]] if best and best[1] >= 94 else []


def _find_group_row(ws, header_row: int, description_row: int | None) -> int | None:
    end = (description_row or header_row) - 1
    best: tuple[float, int] | None = None
    for r in range(1, max(1, end) + 1):
        vals = [str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1) if ws.cell(r, c).value not in (None, "")]
        if not vals:
            continue
        short = sum(1 for v in vals if len(v) <= 35) / len(vals)
        repeated = 1 - (len(set(map(key_norm, vals))) / max(len(vals), 1))
        score = short + repeated
        if best is None or score > best[0]:
            best = (score, r)
    return best[1] if best and best[0] >= 0.75 else None


def _find_requirement_row(ws, description_row: int | None, header_row: int) -> int | None:
    start = (description_row or max(1, header_row - 2)) + 1
    for r in range(start, header_row):
        vals = [key_norm(str(ws.cell(r, c).value or "")) for c in range(1, ws.max_column + 1)]
        if any("optional" in v or "opcional" in v or "required" in v or "obligatorio" in v for v in vals):
            return r
    return None


def _is_required(marker: Any) -> bool:
    n = key_norm(str(marker or ""))
    if "optional" in n or "opcional" in n:
        return False
    return True


def _field_role(field: dict, group: str | None) -> tuple[str, bool, str]:
    label = key_norm(field.get("label") or "")
    desc = key_norm(field.get("description") or "")
    semantic = key_norm((field.get("contract") or {}).get("semantic") or field.get("canonical") or "")
    field_class = field.get("field_class") or "UNKNOWN"
    ext = str(field.get("external_id") or "")
    combined = " ".join([label, desc, semantic, key_norm(group or "")])

    if field_class == "IMAGE":
        return ROLE_MEDIA, True, "La plantilla solicita una URL de imagen de producto."
    if field_class == "SELLER_DATA":
        return ROLE_SELLER, False, "Dato comercial/operativo del vendedor; no se inventa por scraping."

    if ext in {"1", "22"} or any(x in combined for x in ["categoria primaria", "primary category", "condicion del producto", "status of the product"]):
        return ROLE_MARKETPLACE, False, "Dato de catálogo/marketplace definido por el negocio, no por la ficha técnica."
    if ext in {"133815", "133816"} or semantic in {"name cn", "name en", "name_cn", "name_en"}:
        return ROLE_DERIVED, False, "Salida logística derivable a partir de identidad validada/traducción controlada."
    if ext == "3" or "sku padre" in combined or "parent sku" in combined:
        return ROLE_SELLER, False, "Relación de variación administrada por el vendedor/marketplace."

    if semantic in {"mpn", "ean", "upc", "gtin", "model", "brand", "product name", "product_name"}:
        return ROLE_IDENTITY, True, "Identidad técnica usada para validar el producto exacto."

    if field_class in {"SCRAPABLE", "DERIVABLE"}:
        return ROLE_SCRAPE, True, "Dato de producto obtenible o demostrable desde fuentes técnicas validadas."

    if any(x in combined for x in [
        "pais de produccion", "country of production", "color", "modelo", "model", "barcode",
        "bluetooth", "conectividad", "connectivity", "auricular", "headphone", "alto", "ancho", "largo",
        "water", "agua", "autonomia", "battery life", "alimentacion", "power source", "caracteristicas",
        "features", "dimensiones", "dimensions", "serial number", "contenido del paquete", "package contents",
        "garantia del producto", "product warranty", "peso del paquete", "package weight",
    ]):
        return ROLE_SCRAPE, True, "La descripción del Excel define un dato técnico verificable."

    return ROLE_REVIEW, False, "La intención no es suficientemente clara; requiere clasificación antes de escribir."


def analyze_template_contract(template: str) -> dict:
    """Read the marketplace workbook first and convert it into an explicit execution contract.

    This is intentionally upstream of web discovery. The scraper should know what the
    workbook needs before collecting evidence, and seller/commercial fields must never
    become search targets by accident.
    """
    wb = load_workbook(template, data_only=False, read_only=False)
    option_index = _build_option_index(wb)
    sheets: list[dict] = []
    targets: list[TemplateTarget] = []

    for ws in wb.worksheets:
        matrix = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, min(ws.max_row, 24) + 1)]
        info = analyze_matrix(matrix)
        if not info.get("fields"):
            continue
        header_row = info["header_row"]
        description_row = info.get("description_row")
        group_row = _find_group_row(ws, header_row, description_row)
        requirement_row = _find_requirement_row(ws, description_row, header_row)
        sheet_targets: list[dict] = []

        for field in info["fields"]:
            c = field["column"]
            group = str(ws.cell(group_row, c).value).strip() if group_row and ws.cell(group_row, c).value not in (None, "") else None
            marker = ws.cell(requirement_row, c).value if requirement_row else None
            required = _is_required(marker)
            role, scrape, reason = _field_role(field, group)
            contract = field.get("contract") or {}
            target = TemplateTarget(
                sheet=ws.title,
                column=c,
                label=field["label"],
                description=field.get("description"),
                group=group,
                external_id=field.get("external_id"),
                canonical=field.get("canonical"),
                semantic=contract.get("semantic"),
                field_class=field.get("field_class") or "UNKNOWN",
                role=role,
                required=required,
                value_type=contract.get("value_type") or "text",
                options=_find_options(option_index, field["label"]),
                scrape=scrape,
                reason=reason,
            )
            targets.append(target)
            sheet_targets.append(target.to_dict())

        sheets.append({
            "sheet": ws.title,
            "header_row": header_row,
            "description_row": description_row,
            "group_row": group_row,
            "requirement_row": requirement_row,
            "fields": sheet_targets,
        })

    counts: dict[str, int] = {}
    for target in targets:
        counts[target.role] = counts.get(target.role, 0) + 1

    scrape_targets = [t for t in targets if t.scrape and t.role != ROLE_MEDIA]
    media_targets = [t for t in targets if t.role == ROLE_MEDIA]
    return {
        "sheets": sheets,
        "summary": {
            "fields_total": len(targets),
            "roles": counts,
            "scrape_targets": len(scrape_targets),
            "media_slots": len(media_targets),
            "seller_inputs": counts.get(ROLE_SELLER, 0),
            "marketplace_inputs": counts.get(ROLE_MARKETPLACE, 0),
            "review_required": counts.get(ROLE_REVIEW, 0),
        },
        "scrape_semantics": list(dict.fromkeys(
            str(t.semantic or t.canonical or _strip_field_id(t.label)) for t in scrape_targets
        )),
        "media_slots": len(media_targets),
    }
