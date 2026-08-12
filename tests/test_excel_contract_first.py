from pathlib import Path

from openpyxl import Workbook

from product_intelligence.batch import detect_items
from product_intelligence.template_contract import analyze_template_contract


def template_path():
    return Path(__file__).resolve().parents[1] / "examples" / "ProductCreationTemplate_reference.xlsx"


def all_fields(plan):
    return [field for sheet in plan["sheets"] for field in sheet["fields"]]


def test_excel_contract_separates_product_media_and_seller_fields():
    plan = analyze_template_contract(str(template_path()))
    summary = plan["summary"]
    assert summary["fields_total"] >= 40
    assert summary["scrape_targets"] >= 15
    assert summary["media_slots"] >= 1
    assert summary["seller_inputs"] >= 1

    fields = all_fields(plan)
    price = next(field for field in fields if "PriceFalabella" in field["label"])
    image = next(field for field in fields if "Imagen principal" in field["label"])
    assert price["role"] == "SELLER_INPUT"
    assert price["scrape"] is False
    assert image["role"] == "MEDIA_TARGET"
    assert image["scrape"] is True


def test_excel_description_defines_bluetooth_contract(tmp_path):
    path = tmp_path / "bluetooth_contract.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Subir plantilla"
    ws.append(["Especificaciones", "Especificaciones", "Precio"])
    ws.append([
        "Selecciona si el producto cuenta con bluetooth. // Select whether the product has Bluetooth.\n\n- Syntax: One value from the list",
        "Ingresa el modelo del producto. // Enter the product model.\n\n- Syntax: Text",
        "Ingresa el precio del producto en falabella.com.",
    ])
    ws.append([" ", " ", " ( Optional ) "])
    ws.append(["CuentaConBluetooth #1568", "Modelo #32", "PriceFalabella #52"])
    ws.append([None, "ABC-123", None])
    opts = wb.create_sheet("Opciones")
    opts.append(["CuentaConBluetooth #1568", "Dummy"])
    opts.append(["Sí", "A"])
    opts.append(["No", "B"])
    wb.save(path)

    plan = analyze_template_contract(str(path))
    bluetooth = next(field for field in all_fields(plan) if "CuentaConBluetooth" in field["label"])
    assert bluetooth["scrape"] is True
    assert bluetooth["value_type"] == "controlled"
    assert "bluetooth" in str(bluetooth["semantic"]).lower()
    assert {str(x) for x in bluetooth["options"]} >= {"Sí", "No"}


def test_unknown_technical_field_is_scrape_target_by_excel_contract(tmp_path):
    path = tmp_path / "generic_product.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Subir plantilla"
    ws.append(["Especificaciones", "Precio", "Principales"])
    ws.append([
        "Indica la velocidad máxima del producto en km/h. // Enter the maximum speed of the product in km/h.\n\n- Syntax: Text",
        "Ingresa el precio que el cliente debe pagar por el producto.",
        "Ingresa la condición física del producto.",
    ])
    ws.append([" ", " ( Optional ) ", " "])
    ws.append(["VelocidadMaxima #999999", "PriceFalabella #52", "Detalles de la condición del Producto #49"])
    ws.append([None, None, None])
    wb.save(path)

    plan = analyze_template_contract(str(path))
    fields = all_fields(plan)
    speed = next(field for field in fields if "VelocidadMaxima" in field["label"])
    price = next(field for field in fields if "PriceFalabella" in field["label"])
    condition = next(field for field in fields if "condición" in field["label"])

    assert speed["role"] == "SCRAPE_TARGET"
    assert speed["scrape"] is True
    assert "VelocidadMaxima" in plan["scrape_semantics"]
    assert price["role"] == "SELLER_INPUT" and price["scrape"] is False
    assert condition["role"] == "MARKETPLACE_INPUT" and condition["scrape"] is False


def test_seller_sku_is_not_product_identity():
    items = detect_items(str(template_path()))
    assert items
    assert all(item.identity.sku is None for item in items)
