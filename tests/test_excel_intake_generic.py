from __future__ import annotations

from openpyxl import Workbook

from product_intelligence.batch import detect_items
from product_intelligence.excel_intake import analyze_workbook_intake, identity_header_kind


def _save_workbook(tmp_path, rows, title="Productos"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for r, values in enumerate(rows, 1):
        for c, value in enumerate(values, 1):
            ws.cell(r, c).value = value
    path = tmp_path / "products.xlsx"
    wb.save(path)
    return path


def test_one_column_part_number_sheet_is_a_valid_product_table(tmp_path):
    path = _save_workbook(
        tmp_path,
        [
            ["Part Number"],
            ["TE-2128S"],
            ["IPC-S042"],
            ["JBLQ350WLBLKAM"],
        ],
    )

    items = detect_items(str(path))

    assert [(item.row, item.identity.mpn) for item in items] == [
        (2, "TE-2128S"),
        (3, "IPC-S042"),
        (4, "JBLQ350WLBLKAM"),
    ]


def test_identity_header_aliases_are_semantically_normalized():
    expected = {
        "PART NUMBER": "part_number",
        "Part_Number": "part_number",
        "manufacturer partnumber": "part_number",
        "Modelo #32": "model",
        "Model No": "model",
        "EAN/UPC": "gtin",
        "código de barras": "gtin",
        "sku_seller": "sku",
        "Merchant SKU": "sku",
    }
    assert {label: identity_header_kind(label) for label in expected} == expected


def test_ean_upc_only_row_becomes_gtin_identity(tmp_path):
    path = _save_workbook(tmp_path, [["EAN/UPC"], ["7501234567893"]])
    items = detect_items(str(path))
    assert len(items) == 1
    assert items[0].identity.gtin == "7501234567893"


def test_model_code_only_row_is_promoted_to_strong_part_number(tmp_path):
    path = _save_workbook(tmp_path, [["Model No"], ["IPC-S042"]])
    items = detect_items(str(path))
    assert len(items) == 1
    assert items[0].identity.model == "IPC-S042"
    assert items[0].identity.mpn == "IPC-S042"


def test_sku_only_row_is_kept_as_sku_fallback_not_relabelled_as_mpn(tmp_path):
    path = _save_workbook(tmp_path, [["sku_seller"], ["STORE-IPC-S042"]])
    items = detect_items(str(path))
    assert len(items) == 1
    assert items[0].identity.sku == "STORE-IPC-S042"
    assert items[0].identity.mpn is None


def test_auxiliary_sheet_is_rejected_by_evidence_not_by_name(tmp_path):
    wb = Workbook()
    help_ws = wb.active
    help_ws.title = "Cualquier nombre"
    help_ws.append(["Instrucciones"])
    help_ws.append(["Completa los datos requeridos antes de cargar productos."])

    product_ws = wb.create_sheet("Otra hoja cualquiera")
    product_ws.append(["Manufacturer PartNumber"])
    product_ws.append(["TE-2128S"])
    product_ws.append(["IPC-S042"])

    path = tmp_path / "mixed.xlsx"
    wb.save(path)

    result = analyze_workbook_intake(str(path))
    items = detect_items(str(path))

    assert [item.identity.mpn for item in items] == ["TE-2128S", "IPC-S042"]
    audits = {sheet.sheet: sheet for sheet in result.sheets}
    assert audits["Cualquier nombre"].accepted is False
    assert audits["Otra hoja cualquiera"].accepted is True


def test_header_selection_prefers_real_product_rows_over_upper_semantic_noise(tmp_path):
    path = _save_workbook(
        tmp_path,
        [
            ["Modelo", "Marca"],
            ["Esta fila explica cómo completar la plantilla", None],
            [None, None],
            ["Part Number", "Descripción"],
            ["TE-2128S", "Producto uno"],
            ["IPC-S042", "Producto dos"],
        ],
    )

    result = analyze_workbook_intake(str(path))
    items = detect_items(str(path))

    assert result.sheets[0].header_row == 4
    assert [item.identity.mpn for item in items] == ["TE-2128S", "IPC-S042"]
