from openpyxl import Workbook, load_workbook
from product_intelligence.excel_mapper_v8 import fill_excel_v8
from product_intelligence.models import ProductIdentity, ProductRecord


def test_seller_sku_output_equals_part_number(tmp_path):
    template = tmp_path / "in.xlsx"
    output = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["MPN", "SKU vendedor"])
    ws.append(["PN-TEST-001", None])
    wb.save(template)
    record = ProductRecord(identity=ProductIdentity(mpn="PN-TEST-001", brand="Marca", model="Modelo", confidence=1.0, match_level="EXACT"))
    fill_excel_v8(template, output, [record])
    result = load_workbook(output)["Productos"]
    assert result["B2"].value == "PN-TEST-001"
