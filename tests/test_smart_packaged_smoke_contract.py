from pathlib import Path

from openpyxl import load_workbook

from product_intelligence.smart_packaged_smoke import run_smoke


def test_run_desktop_exposes_smart_e2e_without_breaking_historical_tail():
    text = Path("run_desktop.py").read_text(encoding="utf-8")
    assert '"--smart-e2e-smoke"' in text
    assert "product_intelligence.smart_packaged_smoke" in text
    assert text.rstrip().endswith("managed_main()")


def test_smart_packaged_smoke_exercises_runtime_write_barrier_and_physical_excel(tmp_path):
    report = run_smoke(tmp_path)

    assert report["status"] == "PASS"
    assert report["entrypoint"] == "SMART_PACKAGED_E2E"
    assert len(report["scenarios"]) >= 3

    exact = next(row for row in report["scenarios"] if row["scenario"] == "EXACT_SKU_WEB_TO_EXCEL")
    assert exact["status"] == "PASS"
    assert exact["category"] == "GENERAL"
    assert exact["resolved_fields"] == ["color"]
    exact_xlsx = Path(exact["output_excel"])
    assert exact_xlsx.is_file()
    wb = load_workbook(exact_xlsx, data_only=False)
    assert wb["Products"]["B2"].value == "Black"

    fallback = next(row for row in report["scenarios"] if row["scenario"] == "PDF_ZERO_WEB_FALLBACK")
    assert fallback["status"] == "PASS"
    assert fallback["pdf_documents_found"] == 0
    assert fallback["broad_web_calls"] == 0
    assert fallback["targeted_web_calls"] == 1
    assert fallback["resolved_fields"] == ["driver_size"]

    sibling = next(row for row in report["scenarios"] if row["scenario"] == "SIBLING_WRITE_BARRIER")
    assert sibling["status"] == "PASS"
    assert sibling["written_value"] is None
    sibling_xlsx = Path(sibling["output_excel"])
    assert sibling_xlsx.is_file()
    wb2 = load_workbook(sibling_xlsx, data_only=False)
    assert wb2["Products"]["B2"].value is None
