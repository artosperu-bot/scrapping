from pathlib import Path
from openpyxl import load_workbook
from product_intelligence.excel_mapper_v8 import fill_excel_v8
from product_intelligence.models import ProductIdentity, ProductRecord

PRODUCTS=[("JBLQ350WLBLKAM","JBL Quantum 350 Wireless","Quantum 350 Wireless"),("JBLENDURRUN3BTBAM","JBL Endurance Run 3 Wireless","Endurance Run 3 Wireless"),("JBLT530CBLKAM","JBL Tune 530C USB-C","Tune 530C")]

def _record(pn,name,model):
    return ProductRecord(identity=ProductIdentity(brand="JBL",product_name=name,model=model,mpn=pn,confidence=.99,match_level="EXACT",identifiers_confirmed=["mpn"]))

def test_three_headphones_fill_consecutive_template_rows(tmp_path):
    root=Path(__file__).resolve().parents[1];template=root/'examples'/'ProductCreationTemplate_reference.xlsx'
    records=[_record(*x) for x in PRODUCTS];assignments={("Subir plantilla",5+i):r for i,r in enumerate(records)}
    output=tmp_path/'headphones_completed.xlsx';trace=tmp_path/'trace.json'
    report=fill_excel_v8(str(template),str(output),records,overwrite=True,trace_path=str(trace),row_assignments=assignments)
    assert output.exists() and trace.exists();assert report['summary']['written_count']>0
    ws=load_workbook(output,data_only=False)['Subir plantilla']
    names=[str(ws[f"A{row}"].value or "") for row in (5,6,7)]
    assert 'Quantum 350' in names[0];assert 'Endurance Run 3' in names[1];assert 'Tune 530C' in names[2]
    assert ws['M5'].value != '999.999,99';assert ws['N5'].value != '999.999,99'
    assert ws['D5'].value != 'Esto es un párrafo';assert ws['AC5'].value != 'Ej. 80 cm x 45 cm x 63 cm // E.g. 80 cm x 45 cm x 63 cm';assert ws['AD5'].value != '10 m/s'
